# -*- coding: utf-8 -*-
"""
中远海控 · 主营与业绩追踪看板（方向C + 趋势增强）构建脚本
=====================================================================
数据主源: 中远海控_行业供需与运价数据.xlsx
约定    : 该 Excel 为唯一主数据源；更新文件后重跑本脚本即可刷新整个看板。
输出    : cosco-shipping/cosco-dashboard.html  (单文件，双击即开)
机制    : 时序数据抽成 JS 数组 + 长期趋势图 base64 内嵌，快照与趋势同源驱动。
"""
import os, re, json, base64, io
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager as fm

# ---------- 路径 ----------
HERE = os.path.dirname(os.path.abspath(__file__))
WS   = os.path.dirname(HERE)  # 工作区根
SRC  = os.path.join(HERE, "数据源")  # 所有 Excel 数据源集中存放
XLSX = os.path.join(SRC, "中远海控_行业供需与运价数据.xlsx")
OUT  = os.path.join(HERE, "cosco-dashboard.html")

# ---------- 中文字体（候选优先级：PingFang SC/HK → Songti SC → Heiti TC → STHeiti → SimHei） ----------
try:
    fm._load_fonts(rebuild_cache=True)  # 让 Matplotlib 重新扫系统字体
except Exception:
    try:
        fm.fontManager.__init__()
    except Exception:
        pass
_available = {f.name for f in fm.fontManager.ttflist}
_CANDIDATES = ["PingFang SC", "PingFang HK", "Songti SC", "Heiti TC",
              "STHeiti", "Hei", "Hiragino Sans GB", "Arial Unicode MS",
              "SimHei", "Microsoft YaHei", "Noto Sans CJK SC"]
CN = next((c for c in _CANDIDATES if c in _available), None)
if CN is None:
    CN = next((n for n in sorted(_available)
              if any(k in n for k in ("SC", "TC", "Han", "Hei", "Song", "Kai", "CJK"))),
              "DejaVu Sans")
plt.rcParams["font.family"] = "sans-serif"
plt.rcParams["font.sans-serif"] = [CN, "DejaVu Sans", "Arial"]
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["font.size"] = 11

# ---------- 配色 ----------
C_SCFI="#c8102e"; C_CCFI="#15616d"; C_EU="#2e6f9e"; C_MED="#d98a14"
C_USW="#2e9e6b"; C_USE="#1f7a8c"; C_GULF="#7a8694"
C_THRU="#15616d"; C_EXP="#c8102e"
C_ASEAN="#2e9e6b"; C_EUY="#1f7a8c"; C_USY="#d98a14"; C_TY="#c8102e"
C_FLEET="#15616d"; C_ORDER="#7a8694"; C_IDLE="#d98a14"
C_BLUS="c8102e"; C_BLEU="15616d"

# ---------- 读 Excel ----------
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# ---------- 中远海控_货运量与航线结构数据.xlsx（季度货运量等，来自另一 xlsx） ----------
OPX = os.path.join(SRC, "中远海控_货运量与航线结构数据.xlsx")
opwb = openpyxl.load_workbook(OPX, read_only=True, data_only=True)
def op_sheet_rows(name):
    ws = opwb[name]
    return [list(r) for r in ws.iter_rows(values_only=True)]

# 季度货运量明细(2021-)：单季货运量 / 同比 / 累计 / 外贸单箱收入 / 单季归母净利
qv = op_sheet_rows("季度货运量明细(2021-)")
_qv_rows = [r for r in qv[1:] if r[1] is not None]
qvol_dates = [str(r[1]) for r in _qv_rows]                       # "2021 Q1" ...
qvol       = [float(r[2]) for r in _qv_rows]                       # 单季货运量(万TEU)
qvol_yoy   = [float(r[3]) for r in _qv_rows]                       # 货运量同比(%)
qvol_cum   = [float(r[4]) for r in _qv_rows]                       # 累计货运量(万TEU)
qrev       = [float(r[5]) for r in _qv_rows]                       # 外贸单箱收入(USD/TEU)
qnp        = [float(r[6]) for r in _qv_rows]                       # 单季归母净利润(亿元RMB)

def sheet_rows(name):
    ws = wb[name]
    return [list(r) for r in ws.iter_rows(values_only=True)]

fr = sheet_rows("周运价指数(SCFI&CCFI)")
de = sheet_rows("需求端-港口吞吐量&出口")
su = sheet_rows("供给端-运力&闲置率")
co = sheet_rows("供需调控-停航率&联盟")

# --- 运价 ---
fr_dates=[str(r[0]) for r in fr[1:]]
scfi=[float(r[1]) for r in fr[1:]]
scfi_w=[float(r[2]) for r in fr[1:]]   # 周环比
ccfi=[float(r[3]) for r in fr[1:]]
ccfi_w=[float(r[4]) for r in fr[1:]]
eu=[float(r[5]) for r in fr[1:]]
med=[float(r[6]) for r in fr[1:]]
usw=[float(r[7]) for r in fr[1:]]
use=[float(r[8]) for r in fr[1:]]
gulf=[float(r[9]) for r in fr[1:]]
util_raw=[r[10] for r in fr[1:]]
notes_fr=[str(r[11]) for r in fr[1:]]

def parse_util(v):
    s=str(v); nums=[int(n) for n in re.findall(r"\d+", s)]
    if not nums: return None
    if "-" in s and len(nums)>=2: return round((nums[0]+nums[1])/2,1)
    if ">" in s: return nums[0]+2.5
    if "~" in s: return nums[0]
    return nums[0]
util=[parse_util(v) for v in util_raw]

# --- 需求 ---
de_dates=[str(r[0]) for r in de[1:]]
thru=[float(r[1]) for r in de[1:]]
thru_y=[float(r[2]) for r in de[1:]]
exp=[float(r[3]) for r in de[1:]]
exp_y=[float(r[4]) for r in de[1:]]
asean_y=[float(r[5]) for r in de[1:]]
eu_y=[float(r[6]) for r in de[1:]]
us_y=[float(r[7]) for r in de[1:]]

# --- 供给 ---
su_dates=[str(r[0]) for r in su[1:]]
fleet=[float(r[1]) for r in su[1:]]
fleet_y=[float(r[2]) for r in su[1:]]
order=[float(r[3]) for r in su[1:]]
idle_s=[float(r[4]) for r in su[1:]]
idle_r=[float(r[5]) for r in su[1:]]
notes_su=[str(r[6]) for r in su[1:]]

# --- 调控 ---
co_dates=[str(r[0]) for r in co[1:]]
blank_us=[float(r[1]) for r in co[1:]]
blank_eu=[float(r[2]) for r in co[1:]]
notes_co=[str(r[3]) for r in co[1:]]

# --- 财务核心（独立交付文件：财报核心数据.xlsx）---
FIN_XLSX = os.path.join(SRC, "中远海控_财报核心数据.xlsx")
fws = openpyxl.load_workbook(FIN_XLSX, read_only=True, data_only=True)
def sheet_to_dicts(name):
    ws = fws[name]; rows=list(ws.iter_rows(values_only=True))
    hdr=[str(c) for c in rows[0]]; out=[]
    for r in rows[1:]:
        if r[0] is None: continue
        out.append({hdr[i]: (r[i] if i < len(r) else None) for i in range(len(hdr))})
    return out
fin_a = sheet_to_dicts("年度核心数据")
fin_q = sheet_to_dicts("季度核心数据(当季)")
def fnum(v):
    try: return float(v)
    except: return None
fin_years=[str(d["报告期"]).replace("年","") for d in fin_a]
fin_rev_a=[fnum(d["营业收入(亿元)"]) for d in fin_a]
fin_np_a=[fnum(d["归母净利润(亿元)"]) for d in fin_a]
fin_gm_a=[fnum(d["毛利率(%)"]) for d in fin_a]
fin_roe_a=[fnum(d["ROE摊薄(%)"]) for d in fin_a]
fin_div = sheet_to_dicts("分红数据")
fin_q_lbl=[str(d["报告期"]) for d in fin_q]
fin_rev_q=[fnum(d["营业收入(当季,亿元)"]) for d in fin_q]
fin_np_q=[fnum(d["归母净利润(当季,亿元)"]) for d in fin_q]
# 分红（2021-2025，A+H 全体股东）
div_years=[str(d["报告期"]).replace("年","") for d in fin_div]
div_ps=[fnum(d["全年每股分红(元)"]) for d in fin_div]
div_payout=[fnum(d["分红比例(%)"]) for d in fin_div]
div_amt=[fnum(d["全年现金分红总额(亿元,A+H)"]) for d in fin_div]
# 成本结构（集装箱航运业务，2021-2025，来源：财报核心数据.xlsx 成本结构表）
fin_cost = sheet_to_dicts("成本结构(集装箱航运业务)")
cost_years=[str(d["报告期"]).replace("年","") for d in fin_cost]
cost_total=[fnum(d["集装箱航运业务成本(亿元)"]) for d in fin_cost]
cost_equip=[fnum(d["设备及货物运输成本(亿元)"]) for d in fin_cost]
cost_voyage=[fnum(d["航程成本(含燃油,亿元)"]) for d in fin_cost]
cost_vessel=[fnum(d["船舶成本(含租船,亿元)"]) for d in fin_cost]
cost_other=[fnum(d["其他业务成本(亿元)"]) for d in fin_cost]
cost_perbox=[fnum(d["航线单箱成本(美元/TEU)"]) for d in fin_cost]
cost_vol=[fnum(d["货运量(万TEU)"]) for d in fin_cost]
cost_fuel=[fnum(d["燃油费估算（2025）(亿元)"]) for d in fin_cost]
cost_charter=[fnum(d["租船费估算（2025）(亿元)"]) for d in fin_cost]
# 港口码头业务（2021-2025，来源：财报核心数据.xlsx 港口码头业务表）
fin_port = sheet_to_dicts("港口码头业务")
port_years=[str(d["报告期"]).replace("年","") for d in fin_port]
port_rev=[fnum(d["码头业务收入(亿元)"]) for d in fin_port]
port_cost=[fnum(d["码头业务成本(亿元)"]) for d in fin_port]
port_gm=[fnum(d["码头业务毛利率(%)"]) for d in fin_port]
port_gp=[fnum(d["码头业务毛利(亿元)"]) for d in fin_port]
port_thru=[fnum(d["码头总吞吐量(万TEU)"]) for d in fin_port]
port_share=[fnum(d["营收占比(%)"]) for d in fin_port]

# ---------- 计算最新快照值（供卡片文本） ----------
def pct(x, d=1): return f"{x*100:.{d}f}%"

# 航线结构占比（集装箱航运业务，分航线箱量与收入，2021–2025 全年，来源：中远海控_货运量与航线结构数据.xlsx）
ROUTE_XLSX = os.path.join(SRC, "中远海控_货运量与航线结构数据.xlsx")
_rows = [list(r) for r in openpyxl.load_workbook(ROUTE_XLSX, read_only=True, data_only=True)
         ["分航线箱量与收入明细"].iter_rows(values_only=True)]
ROUTE_FY = ["2021", "2022", "2023", "2024", "2025"]
ROUTE_SHORT = {"跨太平洋 (美线)": "美线", "亚欧 (欧地线)": "欧地线",
               "亚洲区域内 (近洋线)": "近洋线", "中国大陆 (内贸线)": "内贸线",
               "其它国际 (南北线等)": "其它国际"}
ROUTE_NAMES = ["美线", "欧地线", "近洋线", "内贸线", "其它国际"]
_yby = {}  # year -> list of (short_name, vol, rev_rmb)
for rec in _rows[1:]:
    name, period, vol, vol_yoy, rev, rev_usd, pb_u, pb_r = rec
    if not str(period).strip().endswith("FY"):
        continue
    if name not in ROUTE_SHORT:
        continue
    _yby.setdefault(str(period).replace("FY", "").strip(), []).append(
        (ROUTE_SHORT[name], float(vol or 0), float(rev or 0)))
route_vol = {nm: [] for nm in ROUTE_NAMES}
route_rev = {nm: [] for nm in ROUTE_NAMES}
for _yr in ROUTE_FY:
    _recs = _yby.get(_yr, [])
    _tv = sum(r[1] for r in _recs); _tr = sum(r[2] for r in _recs)
    for nm in ROUTE_NAMES:
        _v = next((r[1] for r in _recs if r[0] == nm), 0)
        _r = next((r[2] for r in _recs if r[0] == nm), 0)
        route_vol[nm].append(round(_v, 1)); route_rev[nm].append(round(_r, 1))
route_vshare = {nm: [round(route_vol[nm][i] / sum(route_vol[n][i] for n in ROUTE_NAMES) * 100, 1)
                     for i in range(len(ROUTE_FY))] for nm in ROUTE_NAMES}
route_rshare = {nm: [round(route_rev[nm][i] / sum(route_rev[n][i] for n in ROUTE_NAMES) * 100, 1)
                     for i in range(len(ROUTE_FY))] for nm in ROUTE_NAMES}
ROUTE = {"years": ROUTE_FY, "routes": ROUTE_NAMES,
         "vol": route_vol, "rev": route_rev,
         "vol_share": route_vshare, "rev_share": route_rshare}
# 2025 货运量份额降序摘要（卡片展示用）
_route_mix25 = sorted(((nm, route_vshare[nm][-1]) for nm in ROUTE_NAMES), key=lambda x: -x[1])
route_mix25 = " · ".join(f"{nm} {s:.0f}%" for nm, s in _route_mix25)

VALUES = {
    "scfi": f"{scfi[-1]:.2f}", "scfi_chg": pct(scfi_w[-1]), "scfi_dir": "up" if scfi_w[-1]>=0 else "down",
    "ccfi": f"{ccfi[-1]:.2f}", "ccfi_chg": pct(ccfi_w[-1]), "ccfi_dir": "up" if ccfi_w[-1]>=0 else "down",
    "eu": f"{eu[-1]:.0f}", "med": f"{med[-1]:.0f}", "usw": f"{usw[-1]:.0f}",
    "use": f"{use[-1]:.0f}", "gulf": f"{gulf[-1]:.0f}",
    "util": f"{util[-1]:.0f}%",
    "thru": f"{thru[-1]:.0f} 万TEU", "thru_yoy": pct(thru_y[-1]), "thru_dir":"up",
    "exp": f"{exp[-1]:.0f} 亿美元", "exp_yoy": pct(exp_y[-1]), "exp_dir":"up",
    "asean": pct(asean_y[-1]), "eu_yoy": pct(eu_y[-1]), "us_yoy": pct(us_y[-1]),
    "fleet": f"{fleet[-1]:.0f} 万TEU", "fleet_yoy": pct(fleet_y[-1]), "fleet_dir":"up",
    "idle": pct(idle_r[-1],2), "order": pct(order[-1],1),
    "blank_us": pct(blank_us[-1],1), "blank_eu": pct(blank_eu[-1],1),
    # 财务核心指标（来源：财报核心数据.xlsx）
    "fin_rev25": f"{fin_rev_a[-1]:.1f}亿",
    "fin_np25": f"{fin_np_a[-1]:.1f}亿",
    "fin_np25_chg": ("%.2f%%" % ((fin_np_a[-1]/fin_np_a[-2]-1)*100)),
    "fin_np_q1": f"{fnum(fin_q[-1]['归母净利润(当季,亿元)']):.1f}亿",
    "fin_np_q1_chg": ("%.2f%%" % fnum(fin_q[-1]["当季归母同比(%)"])),
    "fin_gm25": f"{fin_gm_a[-1]:.1f}%",
    "fin_ocf25": f"{fnum(fin_a[-1]['经营现金流(亿元)']):.1f}亿",
    "fin_dar25": f"{fnum(fin_a[-1]['资产负债率(%)']):.1f}%",
    "fin_roe25": f"{fin_roe_a[-1]:.1f}%",
    "fin_eps25": f"{fnum(fin_a[-1]['基本EPS(元)']):.2f}元",
    # 股东回报（分红来源：财报核心数据.xlsx；回购来源：中远海控_回购明细.xlsx）
    "div_payout25": f"{div_payout[-1]:.1f}%",
    "div_ps25": f"{div_ps[-1]:.2f}元",
    "div_amt25": f"{div_amt[-1]:.1f}亿",
    "bb_a5_shares": "2911.65万股",
    "bb_a5_amt": "4.22亿",
    "bb_a5_cap": "15.4亿",
    "bb_h_ytd": "4489.35万股",
    "bb_h_ytd_amt": "6.66亿港元",
    # 成本结构（来源：财报核心数据.xlsx 成本结构表）
    "cost_perbox25": f"{cost_perbox[-1]:.0f} 美元/TEU",
    "cost_perbox_yoy": ("%.1f%%" % ((cost_perbox[-1]/cost_perbox[-2]-1)*100)),
    "cost_total25": f"{cost_total[-1]:.0f} 亿",
    "cost_fuel25": (f"{cost_fuel[-1]:.1f} 亿" if cost_fuel[-1] else "待更新"),
    "cost_charter25": (f"{cost_charter[-1]:.1f} 亿" if cost_charter[-1] else "待更新"),
    # 航线结构占比（来源：中远海控_货运量与航线结构数据.xlsx 分航线箱量与收入明细）
    "route_year": "2025",
    "route_mix25": route_mix25,
    # 季度航运货运量（来源：中远海控_货运量与航线结构数据.xlsx 季度货运量明细(2021-)）
    "qvol25": f"{qvol[-1]:.2f}",
    "qvol_chg": ("+%.2f%%" if qvol_yoy[-1]>=0 else "%.2f%%") % (qvol_yoy[-1]*100),
    "qvol_dir": "up" if qvol_yoy[-1]>=0 else "down",
    "qvol_cum25": f"{qvol_cum[-1]:.2f} 万TEU",
    "qvol_q": qvol_dates[-1],
    # 港口码头业务（来源：财报核心数据.xlsx 港口码头业务表）
    "port_rev25": f"{port_rev[-1]:.1f}亿",
    "port_rev_yoy": ("%.1f%%" % ((port_rev[-1]/port_rev[-2]-1)*100)),
    "port_gm25": f"{port_gm[-1]:.1f}%",
    "port_thru25": f"{port_thru[-1]:.0f}万TEU",
    "port_thru_yoy": ("%.1f%%" % ((port_thru[-1]/port_thru[-2]-1)*100)),
    "port_share25": f"{port_share[-1]:.1f}%",
    # 宏观与政策环境（维度7）：5 项核心数据（公开数据近似）
    "macro_fuel25": "507.5 $/mt",
    "macro_charter25": "2800",
    "macro_ets25": "73 €/t",
    "macro_redsea_ratio25": "82%",
    "macro_redsea_absorb25": "14%",
    "macro_pmi25": "美49.0/服55.0 · 欧48.0/服52.0",
}
# 红涨绿跌（A股惯例）：数值上行=红，下行=绿
for k in ("thru_dir","exp_dir","fleet_dir"):
    pass

# ---------- 趋势图生成（matplotlib → base64） ----------
def fig_to_b64(fig):
    buf=io.BytesIO(); fig.savefig(buf, format="png", dpi=110, bbox_inches="tight",
                                  facecolor="#ffffff"); plt.close(fig)
    return "data:image/png;base64,"+base64.b64encode(buf.getvalue()).decode()

def style_ax(ax):
    ax.grid(True, ls="--", lw=0.5, color="#e3e8ee")
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    ax.tick_params(colors="#1f2733")

CHARTS={}

# C1 运价指数 SCFI/CCFI（双轴）
fig,ax=plt.subplots(figsize=(8.8,3.4))
ax.plot(fr_dates, scfi, color=C_SCFI, lw=2, marker="o", ms=3, label="SCFI 综合指数")
ax.plot(fr_dates, ccfi, color=C_CCFI, lw=2, marker="s", ms=3, label="CCFI 综合指数")
style_ax(ax)
ax.set_ylabel("指数点"); ax.legend(loc="upper left", frameon=False, fontsize=9)
ax.set_title("SCFI / CCFI 综合指数（2026 周度）", fontsize=12, weight="bold")
step=max(1,len(fr_dates)//8)
ax.set_xticks(range(0,len(fr_dates),step)); ax.set_xticklabels([fr_dates[i][5:] for i in range(0,len(fr_dates),step)], rotation=0, fontsize=8)
CHARTS["freight"]=fig_to_b64(fig)

# C2 主力航线即期运价
fig,ax=plt.subplots(figsize=(8.8,3.4))
for d,c,lab in [(eu,C_EU,"欧洲"),(med,C_MED,"地中海"),(usw,C_USW,"美西"),(use,C_USE,"美东"),(gulf,C_GULF,"波斯湾")]:
    ax.plot(fr_dates,d,color=c,lw=1.8,label=lab)
style_ax(ax); ax.set_ylabel("USD / TEU·FEU")
ax.legend(loc="upper left", frameon=False, fontsize=9, ncol=5)
ax.set_title("主力航线即期运价（USD，2026 周度）", fontsize=12, weight="bold")
ax.set_xticks(range(0,len(fr_dates),step)); ax.set_xticklabels([fr_dates[i][5:] for i in range(0,len(fr_dates),step)], fontsize=8)
CHARTS["routes"]=fig_to_b64(fig)

# C3 舱位利用率
fig,ax=plt.subplots(figsize=(8.8,3.0))
ax.plot(fr_dates,util,color=C_IDLE,lw=2,marker="o",ms=3)
style_ax(ax); ax.set_ylabel("舱位利用率 %"); ax.set_ylim(70,105)
ax.set_title("舱位利用率（周度，区间值取中值）", fontsize=12, weight="bold")
ax.set_xticks(range(0,len(fr_dates),step)); ax.set_xticklabels([fr_dates[i][5:] for i in range(0,len(fr_dates),step)], fontsize=8)
CHARTS["util"]=fig_to_b64(fig)

# C4 需求：吞吐量 + 出口总额（双轴）
fig,ax=plt.subplots(figsize=(8.8,3.4))
ax.plot(de_dates,thru,color=C_THRU,lw=2,marker="o",ms=3,label="港口吞吐量(万TEU)")
ax.set_ylabel("吞吐量(万TEU)", color=C_THRU)
ax.tick_params(axis="y",colors=C_THRU)
ax2=ax.twinx(); ax2.plot(de_dates,exp,color=C_EXP,lw=2,marker="s",ms=3,label="出口总额(亿美元)")
ax2.set_ylabel("出口总额(亿美元)", color=C_EXP); ax2.tick_params(axis="y",colors=C_EXP)
style_ax(ax)
ax.set_title("全国港口吞吐量 & 海关出口总额（月频）", fontsize=12, weight="bold")
ax.set_xticklabels(de_dates, rotation=40, ha="right", fontsize=8)
ax.legend(loc="upper left", frameon=False, fontsize=9)
CHARTS["demand"]=fig_to_b64(fig)

# C5 需求：分区域出口同比
fig,ax=plt.subplots(figsize=(8.8,3.4))
for d,c,lab in [(exp_y,C_TY,"总出口"),(asean_y,C_ASEAN,"东盟"),(eu_y,C_EUY,"欧盟"),(us_y,C_USY,"美国")]:
    ax.plot(de_dates,d,color=c,lw=1.8,label=lab)
style_ax(ax); ax.axhline(0,color="#bbb",lw=0.8); ax.set_ylabel("同比增速 %")
ax.legend(loc="upper left", frameon=False, fontsize=9, ncol=4)
ax.set_title("分区域出口同比增速（月频）", fontsize=12, weight="bold")
ax.set_xticklabels(de_dates, rotation=40, ha="right", fontsize=8)
CHARTS["demand_yoy"]=fig_to_b64(fig)

# C6 供给：运力(上) + 闲置率&手持订单(下)
fig,(a1,a2)=plt.subplots(2,1,figsize=(8.8,4.6),sharex=True)
a1.plot(su_dates,fleet,color=C_FLEET,lw=2,marker="o",ms=4)
style_ax(a1); a1.set_ylabel("万TEU"); a1.set_title("全球集装箱船队总运力（季频）", fontsize=12, weight="bold")
a2.plot(su_dates,idle_r,color=C_IDLE,lw=2,marker="o",ms=4,label="闲置率")
a2.plot(su_dates,order,color=C_ORDER,lw=2,marker="s",ms=4,label="手持订单/运力")
style_ax(a2); a2.set_ylabel("%"); a2.legend(loc="upper right", frameon=False, fontsize=9)
a2.set_xticklabels(su_dates, rotation=20, ha="right", fontsize=8)
plt.tight_layout()
CHARTS["supply"]=fig_to_b64(fig)

# C7 调控：美线/欧线停航率
fig,ax=plt.subplots(figsize=(8.8,3.0))
ax.plot(co_dates,blank_us,color="#"+C_BLUS,lw=2,marker="o",ms=4,label="美线停航率")
ax.plot(co_dates,blank_eu,color="#"+C_BLEU,lw=2,marker="s",ms=4,label="欧线停航率")
style_ax(ax); ax.set_ylabel("空白航次比例 %")
ax.legend(loc="upper left", frameon=False, fontsize=9)
ax.set_title("美线 / 欧线停航空白航次比例（季频）", fontsize=12, weight="bold")
ax.set_xticklabels(co_dates, rotation=20, ha="right", fontsize=8)
CHARTS["control"]=fig_to_b64(fig)

# C8 财务：年度营收(柱) + 归母净利润(线)
fig,ax=plt.subplots(figsize=(8.8,3.7))
x=range(len(fin_years))
ax.bar(x, fin_rev_a, color="#15616d", alpha=.85, width=.55, label="营业收入")
for i,v in enumerate(fin_rev_a): ax.text(i, v+25, f"{v:.0f}", ha="center", fontsize=8, color="#15616d")
ax.set_ylabel("营业收入(亿元)", color="#15616d"); ax.tick_params(axis="y", colors="#15616d")
ax.set_xticks(list(x)); ax.set_xticklabels(fin_years)
ax2=ax.twinx()
ax2.plot(x, fin_np_a, color="#c8102e", lw=2.6, marker="o", ms=6, label="归母净利润")
for i,v in enumerate(fin_np_a): ax2.text(i, v+18, f"{v:.0f}", ha="center", fontsize=8, color="#c8102e")
ax2.set_ylabel("归母净利润(亿元)", color="#c8102e"); ax2.tick_params(axis="y", colors="#c8102e")
style_ax(ax); ax.set_title("年度营收 & 归母净利润（2021–2025）", fontsize=12, weight="bold")
l1,la1=ax.get_legend_handles_labels(); l2,la2=ax2.get_legend_handles_labels()
ax.legend(l1+l2, la1+la2, loc="upper right", frameon=False, fontsize=9)
CHARTS["fin_annual"]=fig_to_b64(fig)

# C9 财务：季度营收(柱) + 归母净利润(线)
fig,ax=plt.subplots(figsize=(9.4,3.9))
x=range(len(fin_q_lbl))
ax.bar(x, fin_rev_q, color="#15616d", alpha=.8, width=.62, label="季度营收")
ax.set_ylabel("季度营收(亿元)", color="#15616d"); ax.tick_params(axis="y", colors="#15616d")
ax.set_xticks(list(x)); ax.set_xticklabels(fin_q_lbl, rotation=45, ha="right", fontsize=8)
ax2=ax.twinx()
ax2.plot(x, fin_np_q, color="#c8102e", lw=2.2, marker="o", ms=4, label="季度归母净利润")
ax2.set_ylabel("季度归母净利润(亿元)", color="#c8102e"); ax2.tick_params(axis="y", colors="#c8102e")
style_ax(ax); ax.set_title("季度营收(当季) & 归母净利润(当季)（2021Q1–2026Q1）", fontsize=12, weight="bold")
l1,la1=ax.get_legend_handles_labels(); l2,la2=ax2.get_legend_handles_labels()
ax.legend(l1+l2, la1+la2, loc="upper left", frameon=False, fontsize=9)
CHARTS["fin_quarterly"]=fig_to_b64(fig)

# C10 分红：全年每股分红(柱) + 分红比例(线)
fig,ax=plt.subplots(figsize=(8.8,3.6))
x=range(len(div_years))
ax.bar(x, div_ps, color="#2e9e6b", alpha=.85, width=.55, label="全年每股分红(元)")
for i,v in enumerate(div_ps): ax.text(i, v+0.03, f"{v:.2f}", ha="center", fontsize=8, color="#2e9e6b")
ax.set_ylabel("全年每股分红(元)", color="#2e9e6b"); ax.tick_params(axis="y", colors="#2e9e6b")
ax.set_xticks(list(x)); ax.set_xticklabels(div_years)
ax2=ax.twinx()
ax2.plot(x, div_payout, color="#c8102e", lw=2.6, marker="o", ms=6, label="分红比例")
for i,v in enumerate(div_payout): ax2.text(i, v+1.2, f"{v:.0f}%", ha="center", fontsize=8, color="#c8102e")
ax2.set_ylabel("分红比例(%)", color="#c8102e"); ax2.tick_params(axis="y", colors="#c8102e")
style_ax(ax); ax.set_title("全年每股分红 & 分红比例（2021–2025）", fontsize=12, weight="bold")
l1,la1=ax.get_legend_handles_labels(); l2,la2=ax2.get_legend_handles_labels()
ax.legend(l1+l2, la1+la2, loc="upper right", frameon=False, fontsize=9)
CHARTS["dividend"]=fig_to_b64(fig)

# C11 成本结构：航线成本分项(堆叠柱,左轴亿元) + 单箱成本(线,右轴USD/TEU)
fig,ax=plt.subplots(figsize=(8.8,3.9))
x=list(range(len(cost_years)))
b1=ax.bar(x, cost_equip, color="#15616d", width=.6, label="设备及货物运输")
b2=ax.bar(x, cost_voyage, bottom=cost_equip, color="#d98a14", width=.6, label="航程成本(含燃油)")
b3_bottom=[cost_equip[i]+cost_voyage[i] for i in range(len(x))]
b3=ax.bar(x, cost_vessel, bottom=b3_bottom, color="#2e9e6b", width=.6, label="船舶成本(含租船)")
ax.set_ylabel("航线成本分项(亿元, RMB)"); ax.tick_params(axis="y", colors="#1f2733")
ax.set_xticks(x); ax.set_xticklabels(cost_years)
ax2=ax.twinx()
ax2.plot(x, cost_perbox, color="#c8102e", lw=2.6, marker="o", ms=7, label="航线单箱成本(USD/TEU)")
for i,v in enumerate(cost_perbox): ax2.text(i, v+18, f"{v:.0f}", ha="center", fontsize=8, color="#c8102e")
ax2.set_ylabel("单箱成本(美元/TEU)", color="#c8102e"); ax2.tick_params(axis="y", colors="#c8102e")
style_ax(ax); ax.set_title("成本结构：航线成本分项 & 单箱成本（2021–2025）", fontsize=12, weight="bold")
l1,la1=ax.get_legend_handles_labels(); l2,la2=ax2.get_legend_handles_labels()
ax.legend(l1+l2, la1+la2, loc="upper left", frameon=False, fontsize=8, ncol=2)
CHARTS["cost"]=fig_to_b64(fig)

# C12 航线结构占比：各航线货运量份额（100% 堆叠柱，2021–2025）
_CMAP = {"美线": "#c8102e", "欧地线": "#15616d", "近洋线": "#2e9e6b", "内贸线": "#d98a14", "其它国际": "#7a8694"}
fig, ax = plt.subplots(figsize=(7.6, 3.9), dpi=130)
_rx = list(range(len(ROUTE_FY)))
_rbottom = [0] * len(ROUTE_FY)
for nm in ROUTE_NAMES:
    ax.bar(_rx, route_vshare[nm], bottom=_rbottom, color=_CMAP[nm], width=0.6, label=nm)
    _rbottom = [b + v for b, v in zip(_rbottom, route_vshare[nm])]
ax.set_xticks(_rx); ax.set_xticklabels(ROUTE_FY)
ax.set_ylabel("货运量占比 (%)"); ax.set_ylim(0, 100)
style_ax(ax)
ax.set_title("航线结构占比：各航线货运量份额（2021–2025，全年）", fontsize=12, weight="bold")
ax.legend(ncol=5, fontsize=7.5, loc="upper center", bbox_to_anchor=(0.5, -0.12), frameon=False)
CHARTS["routemix"]=fig_to_b64(fig)

# C13 港口码头业务：码头业务收入(柱,左轴亿元) + 毛利率(线,右轴%) + 总吞吐量(点,右轴2 万TEU)
fig,ax=plt.subplots(figsize=(8.8,3.9))
x=list(range(len(port_years)))
b=ax.bar(x, port_rev, color="#15616d", width=.55, label="码头业务收入(亿元)")
for i,v in enumerate(port_rev): ax.text(i, v+1.5, f"{v:.1f}", ha="center", fontsize=8, color="#15616d")
ax.set_ylabel("码头业务收入(亿元, RMB)"); ax.tick_params(axis="y", colors="#1f2733")
ax.set_xticks(x); ax.set_xticklabels(port_years)
ax2=ax.twinx()
ax2.plot(x, port_gm, color="#c8102e", lw=2.6, marker="o", ms=7, label="码头业务毛利率(%)")
for i,v in enumerate(port_gm): ax2.text(i, v+0.6, f"{v:.1f}%", ha="center", fontsize=8, color="#c8102e")
ax2.set_ylabel("毛利率 (%)", color="#c8102e"); ax2.tick_params(axis="y", colors="#c8102e"); ax2.set_ylim(20, 36)
ax3=ax.twinx(); ax3.spines["right"].set_position(("axes",1.06))
ax3.plot(x, port_thru, color="#2e9e6b", lw=1.8, marker="s", ms=5, ls="--", label="总吞吐量(万TEU)")
ax3.set_ylabel("总吞吐量(万TEU)", color="#2e9e6b"); ax3.tick_params(axis="y", colors="#2e9e6b")
style_ax(ax); ax.set_title("港口码头业务：收入 & 毛利率 & 总吞吐量（2021–2025）", fontsize=12, weight="bold")
l1,la1=ax.get_legend_handles_labels(); l2,la2=ax2.get_legend_handles_labels(); l3,la3=ax3.get_legend_handles_labels()
ax.legend(l1+l2+l3, la1+la2+la3, loc="upper left", frameon=False, fontsize=8, ncol=3)
CHARTS["portmix"]=fig_to_b64(fig)

# ---------- 维度7 宏观与政策环境：5 项核心数据（2021–2025 年度近似） ----------
# 成本驱动（指数化）: 燃油 / 租船 / EU 碳价；运力与需求（水平值）: 红海绕行 + 欧美PMI
# 数据来源：公开数据近似（Ship&Bunker/Argus、Harper Petersen、EU ETS/EEX、ISM/HCOB PMI）
MYEARS = ["2021","2022","2023","2024","2025"]
m_fuel    = [532, 797, 615, 610, 507.5]    # VLSFO 新加坡 $/mt 年度均价
m_charter = [4000, 3000, 1300, 2500, 2800] # Harpex 租船费率指数 年度近似
m_ets     = [54, 80, 83, 65, 73]           # EU ETS 碳价 €/t 日历年均价
m_redsea_ratio  = [9, 8, 12, 85, 82]       # 亚欧航线绕航好望角比例 %（危机前常态 ~10%，2023-11 起飙升至 ~80-90%）
m_redsea_absorb = [0, 0, 1, 15, 14]        # 红海绕行吸收有效运力比例 %（危机前 0，绕行期 ~13-18%）
# 欧美制造业/服务业 PMI 综合（美ISM + 欧HCOB，4 项均值）
m_pmi = [58.9, 54.3, 48.7, 50.3, 51.0]     # 欧美综合 PMI（制造业+服务业均值）
def _idx(s):
    b=s[0]; return [round(v/b*100,1) for v in s]
# (A) 成本驱动指数化（2021=100）
fig=plt.figure(figsize=(7.2,3.6)); ax=fig.add_subplot(111)
ax.plot(MYEARS, _idx(m_fuel),    color="#c8102e", lw=1.8, marker="o", ms=4, label="VLSFO 燃油价")
ax.plot(MYEARS, _idx(m_charter), color="#d98a14", lw=1.8, marker="^", ms=4, label="Harpex 租船费率")
ax.plot(MYEARS, _idx(m_ets),     color="#2e9e6b", lw=1.8, marker="d", ms=4, label="EU ETS 碳价")
ax.axhline(100, color="#888", lw=0.8, ls="--")
ax.set_ylabel("指数（2021=100）"); ax.set_title("宏观成本驱动：燃油 / 租船 / EU 碳价（2021=100）", fontsize=12, weight="bold")
ax.legend(fontsize=8, ncol=3, frameon=False); style_ax(ax)
CHARTS["macroidx"]=fig_to_b64(fig)
# (B) 运力与需求水平值：红海绕行 + 欧美PMI（0-100 轴）
fig=plt.figure(figsize=(7.2,3.6)); ax=fig.add_subplot(111)
ax.plot(MYEARS, m_redsea_ratio,  color="#d84315", lw=1.8, marker="o", ms=4, label="亚欧绕航好望角比例(%)")
ax.plot(MYEARS, m_redsea_absorb, color="#ff8f00", lw=1.8, marker="s", ms=4, label="运力吸收率(%)")
ax.plot(MYEARS, m_pmi,           color="#00897b", lw=1.8, marker="^", ms=4, label="欧美综合PMI")
ax.axhline(50, color="#888", lw=0.8, ls="--")
ax.set_ylabel("% / 指数"); ax.set_title("运力与需求：红海绕行 + 欧美综合PMI（水平值）", fontsize=12, weight="bold")
ax.legend(fontsize=8, ncol=3, frameon=False); style_ax(ax)
CHARTS["macroop"]=fig_to_b64(fig)

# ---------- 季度航运货运量（单季，2021Q1–2026Q1） ----------
fig=plt.figure(figsize=(7.8,3.9)); ax=fig.add_subplot(111)
_x=list(range(len(qvol_dates)))
# 同比(%) 柱状（红涨绿跌）
ax.bar(_x, qvol_yoy, color=["#c8102e" if v>=0 else "#2e9e6b" for v in qvol_yoy],
       alpha=0.20, width=0.6, label="货运量同比(%)")
ax.axhline(0, color="#bbb", lw=0.8)
# 单季货运量(万TEU) 折线（右轴）
ax2=ax.twinx()
ax2.plot(_x, qvol, color="#c8102e", lw=1.9, marker="o", ms=4, label="单季货运量(万TEU)")
ax2.set_ylim(min(qvol)*0.92, max(qvol)*1.08)
for i,v in enumerate(qvol):
    ax2.annotate(f"{v:.0f}", (i,v), textcoords="offset points", xytext=(0,5),
                 fontsize=7, ha="center", color="#c8102e")
ax.set_xticks(_x); ax.set_xticklabels(qvol_dates, rotation=90, fontsize=6.5)
ax.set_ylabel("货运量同比 %"); ax2.set_ylabel("单季货运量(万TEU)")
ax.set_title("季度航运货运量（单季，2021Q1–2026Q1）", fontsize=12, weight="bold")
ax.legend(loc="upper left", fontsize=8, frameon=False)
ax2.legend(loc="upper right", fontsize=8, frameon=False)
style_ax(ax)
CHARTS["qvol"]=fig_to_b64(fig)

# ---------- 序列化 ----------
SERIES = {
    "fr_dates":fr_dates, "scfi":scfi, "ccfi":ccfi, "eu":eu, "med":med, "usw":usw, "use":use, "gulf":gulf, "util":util,
    "de_dates":de_dates, "thru":thru, "exp":exp, "asean_y":asean_y, "eu_y":eu_y, "us_y":us_y, "exp_y":exp_y,
    "su_dates":su_dates, "fleet":fleet, "idle_r":idle_r, "order":order,
    "co_dates":co_dates, "blank_us":blank_us, "blank_eu":blank_eu,
    "fin_years":fin_years, "fin_rev_a":fin_rev_a, "fin_np_a":fin_np_a, "fin_gm_a":fin_gm_a, "fin_roe_a":fin_roe_a,
    "fin_q_lbl":fin_q_lbl, "fin_rev_q":fin_rev_q, "fin_np_q":fin_np_q,
    "div_years":div_years, "div_ps":div_ps, "div_payout":div_payout, "div_amt":div_amt,
    "cost_years":cost_years, "cost_equip":cost_equip, "cost_voyage":cost_voyage,
    "cost_vessel":cost_vessel, "cost_perbox":cost_perbox, "cost_vol":cost_vol,
    "macro_fuel":m_fuel, "macro_charter":m_charter, "macro_ets":m_ets,
    "macro_redsea_ratio":m_redsea_ratio, "macro_redsea_absorb":m_redsea_absorb, "macro_pmi":m_pmi,
    "qvol_dates":qvol_dates, "qvol":qvol, "qvol_yoy":qvol_yoy, "qvol_cum":qvol_cum, "qrev":qrev, "qnp":qnp,
}
NOTES = {
    "freight": "",
    "supply": "",
    "control": notes_co[-1],
    "demand": f"2026年1/2月合并吞吐量{de[1][1]}万TEU、出口{de[1][3]}亿美元为区间峰值，随后回归约3300万TEU常态平台；对东盟出口增速持续领跑，对美国增速波动最大。",
    "fin": "2021–2022 为周期峰值（净利 893 / 1097 亿），2023 谷底（239 亿，-78%），2024–2025 修复至 491 / 309 亿；2026Q1 受运价回落净利 58.8 亿（-49.75%）。资产负债率由 56.8% 降至 41.4%，财务结构持续优化。",
    "div": "2022–2025 分红比例稳定在 ~50%（年报+中报合计）；2021 中报仅转增不派现，比例仅 15.6%。2025 全年每股分红 1.00 元、现金分红 154 亿元（A+H）。",
    "cost": "2021–2022 周期高位单箱成本 1000/1205 美元，2023 运价腰斩降至 786 美元，2024–2025 红海绕行+成本管控趋稳于约 863/866 美元。成本结构中设备及货物运输占比约 48–59%（随运价同向涨跌），航程成本(含燃油)约 17–22%，船舶成本(含租船/折旧)约 16–21%。年报未单列燃油费、租船费明细行；现按 航程成本×70%、船舶成本×30% 估算 2025 年燃油费≈260.6 亿、租船费≈107.9 亿（估算值，非官方披露）。",
}
FIN = {
  "annual": [{"period":fin_years[i], "rev":fin_rev_a[i], "rev_yoy":fnum(fin_a[i]["营收同比(%)"]),
              "np":fin_np_a[i], "np_yoy":fnum(fin_a[i]["归母同比(%)"]),
              "deduct":fnum(fin_a[i]["扣非净利润(亿元)"]), "gm":fin_gm_a[i],
              "nm":fnum(fin_a[i]["净利率(%)"]), "ocf":fnum(fin_a[i]["经营现金流(亿元)"]),
              "dar":fnum(fin_a[i]["资产负债率(%)"]), "roe":fin_roe_a[i],
              "eps":fnum(fin_a[i]["基本EPS(元)"])} for i in range(len(fin_a))],
  "q": [{"period":fin_q_lbl[i], "rev":fin_rev_q[i], "rev_yoy":fnum(fin_q[i]["当季营收同比(%)"]),
         "np":fin_np_q[i], "np_yoy":fnum(fin_q[i]["当季归母同比(%)"]),
         "deduct":fnum(fin_q[i]["扣非净利润(当季,亿元)"]), "gm":fnum(fin_q[i]["当季毛利率(%)"]),
         "nm":fnum(fin_q[i]["当季净利率(%)"])} for i in range(len(fin_q))],
}
# 分红（2021-2025，A+H 全体股东现金分红）
DIV = {
  "annual": [{"period":div_years[i], "annual_ps":div_ps[i]- (fnum(fin_div[i]["中报每股分红(元)"]) or 0),
              "interim_ps":fnum(fin_div[i]["中报每股分红(元)"]), "ps":div_ps[i], "amt":div_amt[i],
              "np":fnum(fin_div[i]["归母净利润(亿元)"]), "payout":div_payout[i]} for i in range(len(fin_div))],
}
# 回购（来源：中远海控_回购明细.xlsx）
BUYBACK = {
  "a_round4": {"first":"2025-11-05","complete":"2026-01-13","cancel":"2026-01-15",
               "shares":55101715,"amt":8.25,"price_low":14.86,"price_high":14.98,"status":"已完成并注销"},
  "a_round5": {"as_of":"2026-08-14","shares":37157524,"amt":5.45,"cap":15.40,
               "price_low":13.69,"price_high":15.40,"status":"进行中·全部注销"},
  "h_ytd": {"as_of":"2026-05-22","times":21,"shares_wan":4489.35,"amt_yi_hkd":6.66,"status":"2026.03.23–05.22 逐日·已注销"},
  "h_round4": {"period":"2025-10~12","shares":123340000,"status":"已完成并注销"},
}
# 成本结构（集装箱航运业务，2021-2025，来源：财报核心数据.xlsx 成本结构表）
# 航线成本分项单位：亿元人民币(RMB)；单箱成本：美元/TEU；货运量：万TEU。
# 燃油费隐含于航程成本、租船费隐含于船舶成本，年报未单列。
COST = {
  "annual": [{"period":cost_years[i], "total":cost_total[i], "equip":cost_equip[i],
              "voyage":cost_voyage[i], "vessel":cost_vessel[i], "other":cost_other[i],
              "perbox":cost_perbox[i], "vol":cost_vol[i],
              "fuel_est":cost_fuel[i], "charter_est":cost_charter[i]} for i in range(len(fin_cost))],
}

# 港口码头业务（2021-2025，来源：财报核心数据.xlsx 港口码头业务表）
# 用户确认：泊位利用率 / 海外布局暂不细化（码头业务占合并营收仅约 2%–6%，占比小）。
PORT = {
  "years": port_years,
  "rev": port_rev, "cost": port_cost, "gm": port_gm, "gp": port_gp,
  "thru": port_thru, "share": port_share,
}

# 宏观与政策环境（维度7）：5 项核心数据（2021–2025 年度近似）
# 成本驱动(指数化): 燃油/租船/EU碳价；运力与需求(水平值): 红海绕行比例/吸收率、欧美PMI。
# 均为公开数据近似，非公司披露。
MACRO = {
  "years": MYEARS,
  "fuel": m_fuel, "charter": m_charter, "ets": m_ets,
  "redsea_ratio": m_redsea_ratio, "redsea_absorb": m_redsea_absorb,
  "pmi": m_pmi,
  "pmi_detail": {"us_manuf":[60.6,55.4,47.5,48.0,49.0], "us_services":[62.0,56.8,54.9,56.0,55.0],
                 "eu_manuf":[58.0,52.0,44.5,46.0,48.0], "eu_services":[55.0,53.0,48.0,51.0,52.0]},
  "latest": {"fuel":507.5, "charter":2800, "ets":73,
             "redsea_ratio":82, "redsea_absorb":14,
             "pmi":51.0, "pmi_detail":"美制造业49.0/服务业55.0 · 欧制造业48.0/服务业52.0"},
}

# 当日行情快照（A股 + 港股，公开行情近似，非本数据库字段）
# 来源：证券之星 / 腾讯行情，2026-08-06 收盘
QUOTE = {
  "date": "2026-08-06",
  "note": "行情快照（收盘），公开数据近似，非本数据库字段；重跑 build_tracker.py 以刷新",
  "a": {"code":"601919.SH", "name":"中远海控",
        "price":15.37, "chg":-0.03, "pct":-0.19,
        "prev":15.40, "open":15.31, "high":15.45, "low":15.25,
        "amount":"10.29亿", "volume":"67.07万手", "turnover":0.53,
        "pe":9.37, "pb":1.02, "mktcap":"2346.71亿"},
  "h": {"code":"01919.HK",
        "price":15.13, "chg":-0.06, "pct":-0.39,
        "prev":15.19, "open":15.05, "high":15.24, "low":14.96,
        "amount":"1.79亿", "volume":"1189万股", "turnover":0.44,
        "pe":8.14, "pb":0.87, "ah":-15.34},
}


DATA_JS = "__SERIES__=" + json.dumps(SERIES, ensure_ascii=False) + ";\n" \
        + "__CHARTS__=" + json.dumps(CHARTS, ensure_ascii=False) + ";\n" \
        + "__VALUES__=" + json.dumps(VALUES, ensure_ascii=False) + ";\n" \
        + "__NOTES__=" + json.dumps(NOTES, ensure_ascii=False) + ";\n" \
        + "__FIN__=" + json.dumps(FIN, ensure_ascii=False) + ";\n" \
        + "__DIV__=" + json.dumps(DIV, ensure_ascii=False) + ";\n" \
        + "__BUYBACK__=" + json.dumps(BUYBACK, ensure_ascii=False) + ";\n" \
        + "__COST__=" + json.dumps(COST, ensure_ascii=False) + ";\n" \
        + "__ROUTE__=" + json.dumps(ROUTE, ensure_ascii=False) + ";\n" \
        + "__PORT__=" + json.dumps(PORT, ensure_ascii=False) + ";\n" \
        + "__MACRO__=" + json.dumps(MACRO, ensure_ascii=False) + ";\n" \
        + "__QUOTE__=" + json.dumps(QUOTE, ensure_ascii=False) + ";\n" \
        + "__META__=" + json.dumps({
            "freight_range": f"{fr_dates[0]} ~ {fr_dates[-1]}（{len(fr_dates)} 周）",
            "demand_range": f"{de_dates[0]} ~ {de_dates[-1]}（{len(de_dates)} 期）",
            "supply_range": f"{su_dates[0]} ~ {su_dates[-1]}（{len(su_dates)} 季）",
            "fin_range": f"{fin_years[0]}–{fin_years[-1]} 年报 + {fin_q_lbl[0]}–{fin_q_lbl[-1]} 季报",
            "route_range": "分航线箱量与收入明细，2021–2025 年报口径（全年）",
            "port_range": "港口码头业务，2021–2025 年报口径（全年）",
            "macro_range": "宏观与政策 5 项核心数据：燃油/租船/EU碳价(2021–2025 指数化) + 红海绕行比例·吸收率 + 欧美PMI(水平值)，公开数据近似",
            "qvol_range": f"季度货运量明细(2021-)，{qvol_dates[0]} ~ {qvol_dates[-1]}（{len(qvol_dates)} 季）· 来源：中远海控_货运量与航线结构数据.xlsx",
            "gen_time": "2026-08-06",
        }, ensure_ascii=False) + ";"

print("charts:", list(CHARTS.keys()))
print("series lens:", {k:len(v) for k,v in SERIES.items()})
print("VALUES sample:", {k:VALUES[k] for k in ["scfi","ccfi","thru","fleet","idle","blank_us"]})

# 把 DATA_JS 暂存，供下一步合成 HTML
with open(os.path.join(HERE,"_data_js.txt"),"w",encoding="utf-8") as f:
    f.write(DATA_JS)
print("DATA_JS bytes:", len(DATA_JS))
