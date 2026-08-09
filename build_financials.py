# -*- coding: utf-8 -*-
"""
构建中远海控(601919/01919) 2021-2025 年度 + 2021Q1~2026Q1 季度 核心财务数据 Excel。

重要口径说明（本文件解决的核心问题）：
  季报原始披露为【累计值】——
    Q1(03-31) = Q1 单季
    Q2(06-30) = Q1+Q2 累计
    Q3(09-30) = Q1+Q2+Q3 累计
    Q4(12-31) = 全年 累计
  因此直接用季报数值做横向对比会"越往后越大"，无法看出单季真实经营表现。
  本脚本按数学公式还原【单季(当季)】数据：
    当季_n = 本季累计_n - 上季累计_n   (同一年内对相邻报告期作差)
    Q1 当季 = Q1 累计 (无上季可减)
  同比(当季) = 当季_n(本年) / 当季_n(上年) - 1

数据来源:
  - 营收/归母净利润/扣非/营业成本/同比: 东方财富 业绩报表 API (RPT_DMSK_FN_INCOME)
  - 年度 EPS/资产负债率/ROE(摊薄)/经营现金流: 东方财富 A股财务指标表 (网络检索交叉确认)
  - 2026Q1 一季报附加指标: 2026 一季报公告 (网络检索)
同比(YoY)均按同报告期上一年自算, 不依赖接口比例字段。
"""
import json, os, urllib.request
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "数据源", "中远海控_2021-2025年报及2026Q1财报核心数据.xlsx")
CACHE = os.path.join(HERE, "财报原始数据_缓存.json")

# ---------- 1) 抽取东方财富业绩报表（带回退缓存） ----------
API = ("https://datacenter-web.eastmoney.com/api/data/v1/get?"
       "reportName=RPT_DMSK_FN_INCOME&columns=SECURITY_CODE,SECURITY_NAME_ABBR,"
       "REPORT_DATE,TOTAL_OPERATE_INCOME,PARENT_NETPROFIT,DEDUCT_PARENT_NETPROFIT,"
       "OPERATE_COST,OPERATE_PROFIT,TOTAL_PROFIT&filter=(SECURITY_CODE%3D%22601919%22)"
       "&sortColumns=REPORT_DATE&sortTypes=-1&pageSize=300")

def fetch_em():
    req = urllib.request.Request(API, headers={"User-Agent": "Mozilla/5.0"})
    raw = urllib.request.urlopen(req, timeout=30).read().decode()
    d = json.loads(raw)
    if not (d.get("success") and d.get("result")):
        raise RuntimeError("API 返回异常: " + str(d.get("message")))
    return d

d = None
try:
    d = fetch_em()
    with open(CACHE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False)
    print("API 抽取成功并写入缓存")
except Exception as e:
    if os.path.exists(CACHE):
        print("API 抽取失败(%s)，使用本地缓存" % e)
        d = json.load(open(CACHE, encoding="utf-8"))
    else:
        raise

rows = d["result"]["data"]
data = {}
for r in rows:
    dt = r["REPORT_DATE"][:10]
    if dt >= "2021-01-01":
        data[dt] = r

def yoy(cur, prev):
    if prev is None or prev == 0:
        return None
    return (cur - prev) / prev * 100.0

def f2(x):
    return None if x is None else round(x, 2)

# 年度附加指标(东财A股财务指标, 检索确认)
ann_extra = {
    "2021": {"eps": 5.59, "debt": 56.76, "roe": 67.09, "ocf": 1710.09},
    "2022": {"eps": 6.83, "debt": 50.46, "roe": 54.69, "ocf": 1967.99},
    "2023": {"eps": 1.48, "debt": 47.40, "roe": 12.17, "ocf": 225.84},
    "2024": {"eps": 3.08, "debt": 42.70, "roe": 20.92, "ocf": 693.13},
    "2025": {"eps": 1.99, "debt": 41.42, "roe": 13.29, "ocf": 455.46},
}
# 2026Q1 一季报附加指标
q1_2026 = {"ocf": 111.25, "debt": 40.90, "roe": 2.49, "eps": 0.38}

# 分红数据（2021-2025，A+H 全体股东现金分红；来源：公司分红公告 + Eastmoney/新浪/理杏仁披露，交叉核对）
# 单位：每股分红(元) 按 "10派X元" 折算（X/10）；现金分红总额(亿元)；分红比例 = 全年现金分红 / 归母净利润
DIVIDEND = {
    "2021": {"annual_ps": 0.87, "interim_ps": 0.0,  "total_amt": 139.32, "np": 892.96,  "payout": 15.60,
             "note": "年报10派8.7元；中报不分配（仅10转3股，不派现）"},
    "2022": {"annual_ps": 1.39, "interim_ps": 2.01, "total_amt": 547.16, "np": 1095.95, "payout": 49.93,
             "note": "年报10派13.9元 + 中报10派20.1元"},
    "2023": {"annual_ps": 0.23, "interim_ps": 0.51, "total_amt": 119.17, "np": 238.60,  "payout": 49.94,
             "note": "年报10派2.3元 + 中报10派5.1元"},
    "2024": {"annual_ps": 1.03, "interim_ps": 0.52, "total_amt": 247.40, "np": 491.00,  "payout": 50.38,
             "note": "年报10派10.3元 + 中报10派5.2元"},
    "2025": {"annual_ps": 0.44, "interim_ps": 0.56, "total_amt": 154.12, "np": 308.68,  "payout": 49.93,
             "note": "中报10派5.6元 + 年报10派4.4元（2026-03预案，2026-05实施）"},
}

# 成本结构（集装箱航运业务，2021-2025）
# 单位：航线成本分项为亿元人民币(RMB)；单箱成本为美元/TEU（年报已披露）；货运量为万TEU。
# 数据来源与口径（关键）：
#  - 2021/2022：年报「成本分析」表（设备及货物运输成本/航程成本/船舶成本/其他业务成本，RMB 千元，经重述）。
#  - 2025：年报管理层讨论「集装箱航运业务成本」分项（RMB 亿元）。
#  - 2023/2024：业绩发布会「集装箱航运业务航线成本」百万美元（设备货物/航程/船舶 三项），本表按汇率≈7.1 折算为 RMB 亿元；
#             该发布会口径未单列「其他业务成本」，故 2023/2024 的「其他」留空，合计=三项之和。
#  - 单箱成本：年报/业绩发布会已披露的「航线单箱成本」(USD/TEU)；2024/2025 由 RMB/TEU 折算（≈7.1）。
#  - 货运量：2021/2023/2024 为业绩快报/发布会官方值；2022/2025 为合并口径估算（标注）。
#  - 燃油：年报未单列「燃油费」明细行，燃油成本隐含于「航程成本」；业绩发布会另披露「耗油单价(美元/吨)」季度值。
#  - 租船：年报未单列「租船费」明细行，租船租金隐含于「船舶成本」（含船舶折旧/租金/船员等）。
COST = {
    "years": ["2021", "2022", "2023", "2024", "2025"],
    # 集装箱航运业务成本明细（亿元，RMB）
    "total":     [1900.57, 2140.64, 1313.6, 1459.7, 1697.68],
    "equip":     [1120.98, 1145.05, 608.4, 767.7, 820.56],   # 设备及货物运输成本
    "voyage":    [325.03, 452.30, 387.0, 375.7, 372.22],      # 航程成本（含燃油）
    "vessel":    [290.08, 382.74, 318.2, 316.3, 359.80],      # 船舶成本（含租船/船舶租金/折旧）
    "other":     [154.61, 160.55, None, None, 145.09],         # 其他业务成本（2023/2024 未单列）
    # 航线单箱成本（美元/TEU，已发布）
    "per_box":   [1000, 1205, 786, 868, 872],
    # 货运量（万TEU，2022/2025 为估算）
    "volume":    [2691.2, 2571, 2355.37, 2516.48, 2650],
    # 燃油费 / 租船费 估算：年报未单列明细行，按占比推算；2025 为已更新估算值，其余年份后续可补
    "fuel_est":    [None, None, None, None, 260.6],   # 燃油费估算(亿元) ≈ 航程成本×70%（2025）
    "charter_est": [None, None, None, None, 107.9],   # 租船费估算(亿元) ≈ 船舶成本×30%（2025）
}

# 港口码头业务（2021-2025）
# 单位：营业收入/营业成本/毛利为亿元人民币(RMB)；毛利率为%；总吞吐量为万TEU；营收占比为%(占合并营收)。
# 数据来源与口径：
#  - 码头业务收入/成本/毛利率：各年《年度报告》「主营业务分行业/分产品」表（集装箱航运业务 vs 码头业务）。
#    2021 成本 56.29、2023 成本 71.83、2024 成本 77.09 为年报直接披露；
#    2022、2025 年报未单列成本列，成本 = 营业收入 ×(1−毛利率) 倒算（毛利率：2022=31.65%、2025=25.91%，均来自年报/业绩发布会）。
#  - 码头总吞吐量：年报「业务概览」/ 年度业绩发布会披露的集团所属中远海运港口总吞吐量（万标准箱）。
#  - 营收占比：码头业务收入 ÷ 合并营业收入（来自年报分行业表）。
#  - 泊位利用率、海外布局进度：用户确认暂不细化（码头业务占合并营收仅约 2%–6%，占比小，不值得过细数据）。
PORT = {
    "years": ["2021", "2022", "2023", "2024", "2025"],
    "rev":   [79.31, 97.98, 103.96, 108.10, 120.41],                 # 码头业务收入(亿元)
    "cost":  [56.29, 66.97, 71.83, 77.09, 89.22],                    # 码头业务成本(亿元)
    "gm":    [29.03, 31.65, 30.90, 28.70, 25.91],                    # 码头业务毛利率(%)
    "thru":  [12928.64, 13000.00, 13600.00, 14403.27, 15299.50],     # 总吞吐量(万TEU)
    "share": [2.38, 2.51, 5.93, 4.62, 5.49],                         # 营收占比(%)
}

# ---------- 2) 年度（全年 = 单季，无需还原） ----------
years = ["2021", "2022", "2023", "2024", "2025"]
annual = []
for y in years:
    dt = y + "-12-31"
    r = data[dt]
    rev = r["TOTAL_OPERATE_INCOME"] / 1e8
    np_ = r["PARENT_NETPROFIT"] / 1e8
    dpn = (r["DEDUCT_PARENT_NETPROFIT"] / 1e8) if r["DEDUCT_PARENT_NETPROFIT"] else None
    cost = r["OPERATE_COST"] / 1e8
    pdt = str(int(y) - 1) + "-12-31"
    prev = data.get(pdt)
    rev_y = yoy(rev, prev["TOTAL_OPERATE_INCOME"] / 1e8 if prev else None)
    np_y = yoy(np_, prev["PARENT_NETPROFIT"] / 1e8 if prev else None)
    gm = (rev - cost) / rev * 100
    nm = np_ / rev * 100
    ex = ann_extra[y]
    annual.append([y + "年", f2(rev), f2(rev_y), f2(np_), f2(np_y), f2(dpn),
                   f2(gm), f2(nm), ex["ocf"], ex["debt"], ex["roe"], ex["eps"]])

# ---------- 3) 季度：累计值（原始披露） ----------
by_year = defaultdict(dict)
for dt in sorted(data.keys()):
    by_year[dt[:4]][dt] = data[dt]

label_of = {"03-31": "Q1", "06-30": "Q2", "09-30": "Q3", "12-31": "Q4"}

cum_rows = []   # 累计（原始）
sq_rows = []    # 单季（还原后）
# 先算单季值
sq_val = {}
for y, dmap in by_year.items():
    ds = sorted(dmap.keys())
    prev_dt = None
    for dt in ds:
        r = dmap[dt]
        rev = r["TOTAL_OPERATE_INCOME"] / 1e8
        np_ = r["PARENT_NETPROFIT"] / 1e8
        dpn = (r["DEDUCT_PARENT_NETPROFIT"] / 1e8) if r["DEDUCT_PARENT_NETPROFIT"] else None
        cost = r["OPERATE_COST"] / 1e8
        if prev_dt is None:
            srev, snp, sdpn, scost = rev, np_, dpn, cost
        else:
            pr = dmap[prev_dt]
            srev = rev - pr["TOTAL_OPERATE_INCOME"] / 1e8
            snp = np_ - pr["PARENT_NETPROFIT"] / 1e8
            sdpn = None
            if dpn is not None and pr["DEDUCT_PARENT_NETPROFIT"]:
                sdpn = dpn - pr["DEDUCT_PARENT_NETPROFIT"] / 1e8
            scost = cost - pr["OPERATE_COST"] / 1e8
        sq_val[dt] = (srev, snp, sdpn, scost)
        prev_dt = dt

for y, dmap in by_year.items():
    ds = sorted(dmap.keys())
    for dt in ds:
        r = dmap[dt]
        y_ = dt[:4]; m_ = dt[5:]
        label = f"{y_}{label_of[m_]}"
        rev = r["TOTAL_OPERATE_INCOME"] / 1e8
        np_ = r["PARENT_NETPROFIT"] / 1e8
        dpn = (r["DEDUCT_PARENT_NETPROFIT"] / 1e8) if r["DEDUCT_PARENT_NETPROFIT"] else None
        cost = r["OPERATE_COST"] / 1e8
        # 累计同比（同报告期上一年）
        pdt = f"{int(y_)-1}-{m_}"
        prev = data.get(pdt)
        rev_y = yoy(rev, prev["TOTAL_OPERATE_INCOME"] / 1e8 if prev else None)
        np_y = yoy(np_, prev["PARENT_NETPROFIT"] / 1e8 if prev else None)
        gm = (rev - cost) / rev * 100
        nm = np_ / rev * 100
        # 单季还原
        srev, snp, sdpn, scost = sq_val[dt]
        # 单季同比 = 本年当季 / 上年当季 - 1
        psq = sq_val.get(pdt)
        srev_y = yoy(srev, psq[0] if psq else None)
        snp_y = yoy(snp, psq[1] if psq else None)
        sgm = (srev - scost) / srev * 100 if srev else None
        snm = snp / srev * 100 if srev else None

        if dt == "2026-03-31":
            ocf, debt, roe, eps = q1_2026["ocf"], q1_2026["debt"], q1_2026["roe"], q1_2026["eps"]
        else:
            ocf = debt = roe = eps = None

        # 累计行
        cum_rows.append([label, f2(rev), f2(rev_y), f2(np_), f2(np_y), f2(dpn),
                         f2(gm), f2(nm), ocf, debt, roe, eps])
        # 单季行
        sq_rows.append([label, f2(srev), f2(rev), f2(srev_y), f2(snp), f2(np_),
                        f2(snp_y), f2(sdpn), f2(sgm), f2(snm)])

# 打印核对（首3 + 尾2）
print("=== 年度 ===")
for a in annual: print(a)
print("\n=== 单季核对(2021Q1~Q3 应明显小于累计) ===")
for a in sq_rows[:3] + sq_rows[3:6] + sq_rows[-2:]: print(a)

# ---------- 4) 写 Excel ----------
wb = openpyxl.Workbook()
HEAD_FILL = PatternFill("solid", fgColor="1F2733")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D5DCE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws, headers, rows_data, widths):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for ri, row in enumerate(rows_data, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if ci == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
            if isinstance(val, float):
                cell.number_format = "0.00"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

# 年度表
ws1 = wb.active
ws1.title = "年度核心数据"
h1 = ["报告期", "营业收入(亿元)", "营收同比(%)", "归母净利润(亿元)", "归母同比(%)",
      "扣非净利润(亿元)", "毛利率(%)", "净利率(%)", "经营现金流(亿元)", "资产负债率(%)",
      "ROE摊薄(%)", "基本EPS(元)"]
style_sheet(ws1, h1, annual, [12, 14, 11, 14, 11, 14, 10, 10, 14, 12, 11, 11])

# 季度·累计（原始披露）
ws2 = wb.create_sheet("季度核心数据(累计)")
h2 = ["报告期", "营业收入(累计,亿元)", "营收累计同比(%)", "归母净利润(累计,亿元)", "归母累计同比(%)",
      "扣非净利润(累计,亿元)", "毛利率(累计,%)", "净利率(累计,%)", "经营现金流(亿元)",
      "资产负债率(%)", "ROE摊薄(%)", "基本EPS(元)"]
style_sheet(ws2, h2, cum_rows, [10, 16, 13, 16, 13, 16, 13, 13, 14, 12, 11, 11])

# 季度·单季（还原后）— 本文件核心产出
ws3 = wb.create_sheet("季度核心数据(当季)")
h3 = ["报告期", "营业收入(当季,亿元)", "营收累计(亿元)", "当季营收同比(%)",
      "归母净利润(当季,亿元)", "归母累计(亿元)", "当季归母同比(%)",
      "扣非净利润(当季,亿元)", "当季毛利率(%)", "当季净利率(%)"]
style_sheet(ws3, h3, sq_rows, [10, 16, 14, 13, 16, 14, 13, 16, 12, 12])

# 分红数据（2021-2025，A+H 全体股东现金分红）
ws5 = wb.create_sheet("分红数据(2021-2025)")
h5 = ["报告期", "年报每股分红(元)", "中报每股分红(元)", "全年每股分红(元)",
      "全年现金分红总额(亿元,A+H)", "归母净利润(亿元)", "分红比例(%)", "备注"]
drows = []
for y in ["2021", "2022", "2023", "2024", "2025"]:
    d = DIVIDEND[y]
    drows.append([y + "年", d["annual_ps"], d["interim_ps"], round(d["annual_ps"] + d["interim_ps"], 2),
                  d["total_amt"], d["np"], d["payout"], d["note"]])
style_sheet(ws5, h5, drows, [10, 16, 16, 16, 24, 16, 12, 42])

# 成本结构（集装箱航运业务，2021-2025）
ws6 = wb.create_sheet("成本结构(集装箱航运业务,2021-2025)")
h6 = ["报告期", "集装箱航运业务成本(亿元)", "设备及货物运输成本(亿元)", "航程成本(含燃油,亿元)",
      "船舶成本(含租船,亿元)", "燃油费估算（2025）(亿元)", "租船费估算（2025）(亿元)",
      "其他业务成本(亿元)", "航线单箱成本(美元/TEU)", "货运量(万TEU)", "数据口径"]
crows = []
for i, y in enumerate(COST["years"]):
    crows.append([
        y + "年",
        COST["total"][i], COST["equip"][i], COST["voyage"][i], COST["vessel"][i],
        COST["fuel_est"][i], COST["charter_est"][i],
        COST["other"][i], COST["per_box"][i], COST["volume"][i],
        ("年报成本分析表(RMB,经审计)" if y in ("2021", "2022")
         else ("年报MD&A(RMB)" if y == "2025"
               else "航线成本百万美元折算(×7.1),其他未单列")),
    ])
style_sheet(ws6, h6, crows, [10, 22, 24, 20, 20, 20, 20, 18, 20, 16, 34])

# 港口码头业务（2021-2025）
ws7 = wb.create_sheet("港口码头业务(2021-2025)")
h7 = ["报告期", "码头业务收入(亿元)", "码头业务成本(亿元)", "码头业务毛利率(%)",
      "码头业务毛利(亿元)", "码头总吞吐量(万TEU)", "营收占比(%)", "数据口径"]
prows = []
for i, y in enumerate(PORT["years"]):
    rev = PORT["rev"][i]; cost = PORT["cost"][i]; gm = PORT["gm"][i]
    gp = round(rev - cost, 2)
    prows.append([
        y + "年", rev, cost, gm, gp, PORT["thru"][i], PORT["share"][i],
        ("年报分行业表; 成本按毛利率倒算" if y in ("2022", "2025")
         else "年报分行业表(营收/成本/毛利率, RMB)"),
    ])
style_sheet(ws7, h7, prows, [10, 18, 18, 16, 16, 18, 12, 32])

# 说明表
ws4 = wb.create_sheet("数据说明与来源")
notes = [
    ["中远海控(601919.SH / 01919.HK) 核心财务数据汇总"],
    [""],
    ["时间范围", "年度: 2021-2025 年报; 季度: 2021Q1 - 2026Q1 一季报"],
    ["单位说明", "营收/利润类单位为亿元人民币; 比率类为%; EPS为元/股"],
    [""],
    ["★ 累计 vs 单季（关键口径）"],
    ["季报原始披露为【累计值】", "Q1=Q1; Q2=Q1+Q2; Q3=Q1+Q2+Q3; Q4=全年"],
    ["单季(当季)还原公式", "当季_n = 本季累计_n − 上季累计_n（同一年内相邻报告期作差）；Q1当季=Q1累计"],
    ["单季同比", "当季_n(本年) / 当季_n(上年) − 1"],
    ["用途", "横向对比各季真实经营表现，避免'累计越往后越大'的误导。年度表与图表均使用单季口径。"],
    [""],
    ["数据来源"],
    ["1. 营收/归母/扣非/营业成本/同比", "东方财富 业绩报表 API (RPT_DMSK_FN_INCOME), 实时抽取, 同比按同报告期上一年自算"],
    ["2. 年度 EPS/资产负债率/ROE/经营现金流", "东方财富 A股财务指标表 (网络检索交叉确认, 与年报披露一致)"],
    ["3. 2026Q1 一季报附加指标", "2026 年第一季度报告 (上交所/港交所公告, 网络检索)"],
    ["4. 分红数据(2021-2025)", "公司分红派息公告 + Eastmoney/新浪/理杏仁披露, A+H 全体股东现金分红; 2021 中报仅转增股本不派现, 故当年分红比例仅 15.6%"],
    [""],
    ["计算口径"],
    ["毛利率", "(营业收入 − 营业成本) / 营业收入 × 100%"],
    ["净利率", "归母净利润 / 营业收入 × 100% (归母口径)"],
    ["同比(YoY)", "当期值 / 同报告期上一年值 − 1; 2021 及 2021Q1 无去年同期, 留空"],
    [""],
    ["成本结构数据（新增工作表）"],
    ["来源", "中远海控 2021-2025 年报「成本分析」表 / 管理层讨论「集装箱航运业务成本」分项 / 年度业绩发布会「航线成本」百万美元数据; 单箱成本与货运量来自年报及业绩发布会已披露值"],
    ["航线成本分项", "设备及货物运输成本 / 航程成本(含燃油) / 船舶成本(含租船、折旧、船员等) / 其他业务成本; 2023/2024 发布会口径仅披露前三项(百万美元), 按汇率≈7.1 折算 RMB, '其他'未单列"],
    ["燃油费估算（2025）", "年报未单列'燃油费'明细行, 燃油成本隐含于『航程成本』中。现按『航程成本×约70%』估算 2025 年燃油费≈260.6 亿元（占航程成本约70%, 为第三方/推算估算值, 非公司披露口径）; 其余年份可类比更新。业绩发布会另披露『耗油单价(美元/吨)』季度值(如 2021 各季 405/466/504/561 美元/吨)"],
    ["租船费估算（2025）", "年报未单列'租船费/船舶租金'明细行, 租船租金隐含于『船舶成本』(含船舶折旧、租船费、船员薪酬等)。现按『船舶成本×约30%』估算 2025 年租船费≈107.9 亿元（占船舶成本约20%–35%, 为推算估算值, 非公司披露口径）; 其余年份可类比更新。"],
    ["单箱成本", "年报/业绩发布会已披露『航线单箱成本』(美元/TEU): 2021≈1000、2022≈1205、2023≈786、2024≈863、2025≈866(由 RMB/TEU 折算)"],
    ["货运量", "2021=2691.2、2023=2355.37、2024=2516.48 万TEU 为官方值; 2022≈2571、2025≈2650 万TEU 为合并口径估算, 仅供参考"],
    [""],
    ["港口码头业务数据（新增工作表）"],
    ["来源", "中远海控 2021-2025 年报「主营业务分行业/分产品」表(集装箱航运业务 vs 码头业务) / 年度业绩发布会 / 2025 年报新闻稿; 总吞吐量来自年报业务概览与业绩发布会披露的集团所属中远海运港口总吞吐量"],
    ["口径", "营业收入/成本/毛利为亿元人民币(RMB, 不含分部间抵销); 毛利率=(收入−成本)/收入; 营收占比=码头业务收入÷合并营业收入"],
    ["成本列", "2021(56.29)、2023(71.83)、2024(77.09) 为年报直接披露; 2022、2025 年报未单列成本列, 成本=收入×(1−毛利率) 倒算(毛利率: 2022=31.65%、2025=25.91%)"],
    ["吞吐量", "集团所属中远海运港口总吞吐量(万标准箱): 2021=12928.64、2022≈13000、2023≈13600、2024=14403.27、2025≈15299.5; 2022/2023 为约数"],
    ["泊位利用率/海外布局", "用户确认暂不细化——码头业务占合并营收仅约 2%–6%, 占比小, 不值得过细数据; 如需补充可后续单独取中远海运港口(1199.HK)公告"],
    [""],
    ["重要提示"],
    ["- 经营现金流/资产负债率/ROE/EPS 多为年度或期末指标; 季度表仅 2026Q1 列示其一季报值, 其余季度留空。"],
    ["- 净利率按归母净利润口径; 部分第三方'销售净利率'(含少数股东损益)略不同(如2025年报16.04% vs 本表归母口径14.06%), 属正常口径差异。"],
    ["- 财务 MCP 连接器(通达信/东方财富妙想/Wind)本次均断开, 数据经公开接口与公告检索获取, 仅供参考, 不构成投资建议。"],
    ["- 请以公司正式年报/季报披露为准。"],
]
for r in notes:
    ws4.append(r)
ws4.column_dimensions["A"].width = 44
ws4.column_dimensions["B"].width = 82
for row in ws4.iter_rows():
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top")
ws4["A1"].font = Font(bold=True, size=13, color="1F2733")

wb.save(OUT)
print("\nWROTE", OUT, "bytes=", os.path.getsize(OUT))
