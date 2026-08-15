# -*- coding: utf-8 -*-
"""
卡片4「财务表现」数据获取 + Excel 存储 + 趋势图 + HTML 更新
-----------------------------------------------------------------
- 数据：A股财报(akshare 新浪接口) 2020至今，季+年
- Excel：写入「财务表现」sheet（聚焦卡片4指标）
- 图表：两张 SVG 趋势图（①营收+归母净利 ②负债率+ROE+每股净资产）
- HTML：更新卡片4 为最新报告期(2026Q1)数据，并嵌入趋势图
"""
import akshare as ak
import pandas as pd
import numpy as np
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
import re, os, json

# 重试会话，规避 TLS 抖动
_session = requests.Session()
_retry = Retry(total=5, backoff_factor=1.0,
               status_forcelist=[429, 500, 502, 503, 504],
               allowed_methods=frozenset(["GET", "POST"]))
for _p in ("https://", "http://"):
    _session.mount(_p, HTTPAdapter(max_retries=_retry))
ak.requests = _session
requests.get = _session.get

STOCK = "sh601898"
EXCEL = "数据源/中煤能源_业务追踪数据.xlsx"
HTML = "中煤能源业务追踪监测清单.html"
KEEP_FROM = "20200101"
SHARES = 13_250_000_000.0  # 中煤能源股本

# ---------- 抓取三表 ----------
income = ak.stock_financial_report_sina(stock=STOCK, symbol="利润表")
balance = ak.stock_financial_report_sina(stock=STOCK, symbol="资产负债表")
cashflow = ak.stock_financial_report_sina(stock=STOCK, symbol="现金流量表")

def keep(df):
    if "类型" in df.columns:
        df = df[df["类型"].astype(str).str.contains("合并", na=False)]
    df = df.copy()
    df["报告日"] = df["报告日"].astype(str)
    return df.sort_values("报告日").reset_index(drop=True)

income, balance, cashflow = keep(income), keep(balance), keep(cashflow)
income = income[income["报告日"] >= KEEP_FROM].reset_index(drop=True)
balance = balance[balance["报告日"] >= KEEP_FROM].reset_index(drop=True)
cashflow = cashflow[cashflow["报告日"] >= KEEP_FROM].reset_index(drop=True)

def pt(r):
    return "年度" if r.endswith("1231") else "季度"

# ---------- 财务表现 指标 ----------
fin = pd.DataFrame()
fin["报告日"] = income["报告日"].astype(str)
fin["期间"] = fin["报告日"].map(pt)
fin["营业收入(亿元)"] = income["营业收入"].astype(float) / 1e8
fin["归母净利润(亿元)"] = income["归属于母公司所有者的净利润"].astype(float) / 1e8
fin["经营现金流净额(亿元)"] = cashflow["经营活动产生的现金流量净额"].astype(float) / 1e8
fin["EPS基本(元)"] = income["基本每股收益"].astype(float)
fin["资产总计(亿元)"] = balance["资产总计"].astype(float) / 1e8
fin["负债合计(亿元)"] = balance["负债合计"].astype(float) / 1e8
fin["归母股东权益(亿元)"] = balance["归属于母公司股东权益合计"].astype(float) / 1e8
fin["资产负债率(%)"] = fin["负债合计(亿元)"] / fin["资产总计(亿元)"] * 100
fin["每股净资产(元)"] = fin["归母股东权益(亿元)"] * 1e8 / SHARES
fin["归母净利率(%)"] = fin["归母净利润(亿元)"] / fin["营业收入(亿元)"] * 100
fin["ROE累计(%)"] = fin["归母净利润(亿元)"] * 1e8 / (fin["归母股东权益(亿元)"] * 1e8) * 100

# 分红率：财报三表不含，标注"见分红公告"，此处留空
fin["分红率(%)"] = np.nan

# 四舍五入
for c in ["营业收入(亿元)","归母净利润(亿元)","经营现金流净额(亿元)","资产负债率(%)",
          "每股净资产(元)","归母净利率(%)","ROE累计(%)"]:
    fin[c] = fin[c].round(2)
fin["EPS基本(元)"] = fin["EPS基本(元)"].round(2)

# ---------- 季度单季还原 ----------
# 财报季报为「年初至报告期累计」值，需还原成真实单季值：
#   同一年内：单季 = 本期累计 - 上一报告期累计；Q1 = Q1累计（直接）。
#   年末 Q4 = 全年累计 - Q3累计。
# 流量指标(累计)需还原：营业收入/归母净利润/经营现金流净额/归母净利率
# 存量指标(时点)不还原：资产负债率/每股净资产
# ROE累计为累计率，单季 ROE 用「单季归母 / 期初期末平均归母权益」估算
def quarter_label(r):
    return f"{r[:4]}Q{int(r[4:6])//3}"

# 季度图包含 Q1~Q4（1231 既是 Q4 季报，也等于全年累计，季度趋势需保留）
fin_q = fin[fin["报告日"].str.endswith(("0331", "0630", "0930", "1231"))].copy().reset_index(drop=True)
# 辅助：按年分组，按报告日顺序计算单季
flow_cols = ["营业收入(亿元)", "归母净利润(亿元)", "经营现金流净额(亿元)", "归母净利率(%)", "EPS基本(元)"]
fin_q_single = fin_q.copy()
prev_by_year = {}
for i, row in fin_q.iterrows():
    y = row["报告日"][:4]
    cum = {c: row[c] for c in flow_cols}
    if row["报告日"].endswith("0331"):
        # Q1 直接为单季
        single = cum
    else:
        prev = prev_by_year.get(y)
        single = {c: (row[c] - prev[c]) for c in flow_cols} if prev else cum
    for c in flow_cols:
        fin_q_single.at[i, c] = round(single[c], 2)
    # ROE单季估算：单季归母 / 期初期末平均归母权益
    eq = row["归母股东权益(亿元)"]
    if row["报告日"].endswith("0331"):
        eq_prev = eq  # 近似用期初=期末
    else:
        eq_prev = prev_by_year.get(y, {}).get("eq", eq)
    roe_single = single["归母净利润(亿元)"] * 1e8 / ((eq + eq_prev) / 2 * 1e8) * 100
    fin_q_single.at[i, "ROE累计(%)"] = round(roe_single, 2)
    # 资产负债率和每股净资产为时点数，沿用（已在 fin 中）
    prev_by_year[y] = {"营业收入(亿元)": row["营业收入(亿元)"],
                       "归母净利润(亿元)": row["归母净利润(亿元)"],
                       "经营现金流净额(亿元)": row["经营现金流净额(亿元)"],
                       "归母净利率(%)": row["归母净利率(%)"],
                       "EPS基本(元)": row["EPS基本(元)"],
                       "eq": eq}
# 给单季表加标签：同一年内只有 Q1 前显示年份（如 2020Q1 Q2 Q3 Q4 2021Q1...）
def quarter_label_smart(dates):
    labels = []
    prev_year = None
    for d in dates:
        y = d[:4]
        q = f"Q{int(d[4:6])//3}"
        if y != prev_year:
            labels.append(f"{y}{q}")
            prev_year = y
        else:
            labels.append(q)
    return labels

fin_q_single["标签"] = quarter_label_smart(fin_q_single["报告日"].tolist())
fin_y = fin[fin["期间"] == "年度"].copy().reset_index(drop=True)
fin_y["标签"] = fin_y["报告日"].str[:4]

# ---------- 写入 Excel：财务表现 sheet ----------
wb = load_workbook(EXCEL)
if "财务表现" in wb.sheetnames:
    del wb["财务表现"]
ws = wb.create_sheet("财务表现")
ws.cell(row=1, column=1, value="卡片4·财务表现（合并口径·财报来源·单位:亿元/%；季度值为真实单季，已做累计还原）").font = \
    __import__("openpyxl").styles.Font(bold=True, size=13, color="1E3A5F")
hdr = ["报告日","期间","营业收入(亿元)","归母净利润(亿元)","经营现金流净额(亿元)",
       "EPS基本(元)","每股净资产(元)","资产负债率(%)","归母净利率(%)","ROE累计(%)","分红率(%)"]
for j, h in enumerate(hdr, 1):
    c = ws.cell(row=3, column=j, value=h)
    c.fill = __import__("openpyxl").styles.PatternFill("solid", fgColor="1E3A5F")
    c.font = __import__("openpyxl").styles.Font(color="FFFFFF", bold=True)
    c.alignment = __import__("openpyxl").styles.Alignment(horizontal="center", wrap_text=True)
# 季度行展示真实单季值（不含 1231，避免与年度全年值重复），年度行展示全年累计值
fin_out = pd.concat([
    fin_q_single[~fin_q_single["报告日"].str.endswith("1231")][hdr].copy(),
    fin_y[hdr].copy()
], ignore_index=True).sort_values("报告日").reset_index(drop=True)
for i, (_, row) in enumerate(fin_out.iterrows(), start=4):
    for j, h in enumerate(hdr, 1):
        v = row[h]
        if isinstance(v, float) and np.isnan(v):
            v = None
        c = ws.cell(row=i, column=j, value=v)
        c.alignment = __import__("openpyxl").styles.Alignment(horizontal="center")
for j in range(1, len(hdr)+1):
    ws.column_dimensions[__import__("openpyxl").utils.get_column_letter(j)].width = 15
ws.freeze_panes = "A4"
wb.save(EXCEL)
print(f"✅ Excel 写入『财务表现』sheet：{len(fin)} 行")

# ---------- 生成 SVG 趋势图 ----------
dates = fin["报告日"].tolist()
labels = [d if d.endswith("1231") else d[4:6] for d in dates]  # 年度显全，季度显月

def build_svg_base(series_list, title, yunit, fname, W=1000, H=420, xlabels=None, label_step=None):
    """series_list: list of (name, values, color). xlabels: 与 values 等长的报告日列表。返回 SVG 字符串"""
    ml, mr, mt, mb = 58, 24, 42, 44
    pw, ph = W - ml - mr, H - mt - mb
    n = len(series_list[0][1]) if series_list else 1
    xs = [ml + (pw * i / (n - 1)) for i in range(n)]
    svg = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" font-family="-apple-system,Segoe UI,PingFang SC,sans-serif">']
    svg.append(f'<rect width="{W}" height="{H}" fill="#ffffff"/>')
    svg.append(f'<text x="{ml}" y="26" font-size="16" font-weight="700" fill="#1e3a5f">{title}</text>')
    # y 轴范围
    allv = [v for _, vals, _ in series_list for v in vals if v is not None]
    ymin, ymax = min(allv), max(allv)
    pad = (ymax - ymin) * 0.15 or 1
    ymin, ymax = ymin - pad, ymax + pad
    def ypos(v):
        return mt + ph - (v - ymin) / (ymax - ymin) * ph
    # 网格 + y 刻度
    for k in range(5):
        yv = ymin + (ymax - ymin) * k / 4
        yy = ypos(yv)
        svg.append(f'<line x1="{ml}" y1="{yy:.1f}" x2="{ml+pw}" y2="{yy:.1f}" stroke="#eef1f5"/>')
        svg.append(f'<text x="{ml-8}" y="{yy+4:.1f}" font-size="11" fill="#9ca3af" text-anchor="end">{yv:.1f}</text>')
    # 轴
    svg.append(f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt+ph}" stroke="#d1d5db"/>')
    svg.append(f'<line x1="{ml}" y1="{mt+ph}" x2="{ml+pw}" y2="{mt+ph}" stroke="#d1d5db"/>')
    # x 标签
    step = label_step if label_step else max(1, n // 10)
    for i in range(n):
        if i % step == 0 or i == n - 1:
            _lab = (xlabels[i] if xlabels else dates[i])
            svg.append(f'<text x="{xs[i]:.1f}" y="{mt+ph+18:.1f}" font-size="10.5" fill="#6b7280" text-anchor="middle">{_lab}</text>')
    # 折线 + 数据点
    for name, vals, color in series_list:
        pts = [(xs[i], ypos(vals[i])) for i in range(n) if vals[i] is not None]
        if len(pts) > 1:
            d = "M" + " L".join(f"{x:.1f} {y:.1f}" for x, y in pts)
            svg.append(f'<path d="{d}" fill="none" stroke="{color}" stroke-width="2.5"/>')
        for x, y in pts:
            svg.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{color}" stroke="#fff" stroke-width="1"/>')
    # 图例
    lx = ml + 4
    for name, _, color in series_list:
        svg.append(f'<rect x="{lx}" y="{H-22}" width="12" height="12" rx="2" fill="{color}"/>')
        svg.append(f'<text x="{lx+17}" y="{H-12}" font-size="11.5" fill="#374151">{name}</text>')
        lx += 18 + len(name) * 12
    svg.append(f'<text x="{ml+pw}" y="{H-12}" font-size="10.5" fill="#9ca3af" text-anchor="end">单位:{yunit}</text>')
    svg.append('</svg>')
    s = "".join(svg)
    with open(fname, "w", encoding="utf-8") as f:
        f.write(s)
    print(f"✅ 生成图表：{fname}")
    return s

def build_sparkline(vals, color, fname):
    """生成卡片内 mini 折线 sparkline SVG，优先使用单季/单期真实数据"""
    W, H = 62, 28
    n = len(vals)
    if n < 2:
        d = f"M2 {H/2:.1f} L{W-2} {H/2:.1f}"
    else:
        clean = [v for v in vals if v is not None]
        ymin, ymax = min(clean), max(clean)
        pad = (ymax - ymin) * 0.1 or 1
        ymin, ymax = ymin - pad, ymax + pad
        pts = []
        for i, v in enumerate(vals):
            if v is None:
                continue
            x = 2 + (W - 4) * i / (n - 1)
            y = H - 2 - (v - ymin) / (ymax - ymin) * (H - 4)
            pts.append(f"{x:.1f} {y:.1f}")
        d = "M" + " L".join(pts)
    svg = (f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" class="spark">'
           f'<path d="{d}" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/></svg>')
    with open(fname, "w", encoding="utf-8") as f:
        f.write(svg)
    print(f"✅ 生成 sparkline：{fname}")
    return svg

# ① 组合大图（长期趋势专区折叠面板使用）
# 季度点用真实单季值、年度点用全年值（避免把累计值当季趋势）
single_map = {r["报告日"]: r for _, r in fin_q_single.iterrows()}
year_map = {r["报告日"]: r for _, r in fin_y.iterrows()}
def mix(col_single, col_year):
    out = []
    for d in dates:
        if d.endswith("1231"):
            out.append(year_map.get(d, {}).get(col_year))
        else:
            out.append(single_map.get(d, {}).get(col_single))
    return out
q_labels_comb = quarter_label_smart([d for d in dates if not d.endswith("1231")])
build_svg_base(
    [("营业收入", mix("营业收入(亿元)", "营业收入(亿元)"), "#c53030"),
     ("归母净利润", mix("归母净利润(亿元)", "归母净利润(亿元)"), "#1e3a5f")],
    "营业收入 vs 归母净利润（季度·单季 / 年度·全年）", "亿元",
    "trend_revenue_profit.svg", W=1000, H=420,
    xlabels=[d[-4:] if d.endswith("1231") else q_labels_comb.pop(0) for d in dates])

q_labels_comb2 = quarter_label_smart([d for d in dates if not d.endswith("1231")])
build_svg_base(
    [("资产负债率", mix("资产负债率(%)", "资产负债率(%)"), "#047857"),
     ("ROE累计", [single_map.get(d, {}).get("ROE累计(%)") if not d.endswith("1231") else year_map.get(d, {}).get("ROE累计(%)") for d in dates], "#b45309"),
     ("每股净资产", mix("每股净资产(元)", "每股净资产(元)"), "#1e3a5f")],
    "资产负债率 / ROE / 每股净资产", "%·元",
    "trend_ratios.svg", W=1000, H=420,
    xlabels=[d[-4:] if d.endswith("1231") else q_labels_comb2.pop(0) for d in dates])

# ② 每个指标单独生成「季度版」与「年度版」两张图（弹窗内可切换）
CHARTS = {}  # key -> {title, q, y}
for key, title, col, unit, color in [
    ("revenue", "营业收入", "营业收入(亿元)", "亿元", "#c53030"),
    ("profit", "归母净利润", "归母净利润(亿元)", "亿元", "#1e3a5f"),
    ("cash", "经营现金流净额", "经营现金流净额(亿元)", "亿元", "#d97706"),
    ("debt", "资产负债率", "资产负债率(%)", "%", "#047857"),
    ("roe", "ROE累计", "ROE累计(%)", "%", "#b45309"),
    ("bps", "每股净资产", "每股净资产(元)", "元", "#2563eb"),
]:
    fq = f"trend_{key}_q.svg"
    fy = f"trend_{key}_y.svg"
    # 季度：用还原后的真实单季值
    fq_df = fin_q_single
    build_svg_base([(title, fq_df[col].tolist(), color)],
                   f"{title}趋势（季度·单季）", unit, fq, W=1000, H=420,
                   xlabels=fq_df["标签"].tolist(), label_step=1)
    # 年度：年报本就是全年值（非累计叠加），直接用
    fy_df = fin_y
    build_svg_base([(title, fy_df[col].tolist(), color)],
                   f"{title}趋势（年度）", unit, fy, W=1000, H=420,
                   xlabels=fy_df["标签"].tolist())
    CHARTS[key] = {"title": title, "q": fq, "y": fy}

# ---------- 生成卡片内 mini sparkline ----------
# 优先使用单季数据；存量指标(负债率/BPS)和年度-only 数据用原序列
SPARKLINES = {
    "revenue": (fin_q_single["营业收入(亿元)"].tolist(), "#c53030"),
    "profit":  (fin_q_single["归母净利润(亿元)"].tolist(), "#c53030"),
    "cash":    (fin_q_single["经营现金流净额(亿元)"].tolist(), "#047857"),
    "debt":    (fin_q_single["资产负债率(%)"].tolist(), "#047857"),
    "roe":     (fin_q_single["ROE累计(%)"].tolist(), "#1e3a5f"),
    "bps":     (fin_q_single["每股净资产(元)"].tolist(), "#1e3a5f"),
}
for key, (vals, color) in SPARKLINES.items():
    build_sparkline(vals, color, f"spark_{key}.svg")

# ---------- 更新 HTML 卡片4 ----------
with open(HTML, encoding="utf-8") as f:
    html = f.read()

latest = fin.iloc[-1]
latest_date = latest["报告日"]
rev = latest["营业收入(亿元)"]; np_ = latest["归母净利润(亿元)"]
ocf = latest["经营现金流净额(亿元)"]; dar = latest["资产负债率(%)"]
roe = latest["ROE累计(%)"]; bvps = latest["每股净资产(元)"]
# 同比（去年同期，取上一年的同报告日或上一行）
prev_year = str(int(latest_date[:4]) - 1) + latest_date[4:]
prev = fin[fin["报告日"] == prev_year]
if prev.empty and latest["期间"] == "季度":
    prev = fin[fin["报告日"] == fin["报告日"].iloc[-2]]  # 退化为上一报告期
if not prev.empty:
    p = prev.iloc[0]
    def pct_delta(cur, base):
        if base in (0, None) or (isinstance(base, float) and np.isnan(base)):
            return ""
        d = (cur - base) / base * 100
        return f'{d:+.1f}%'
    rev_d = pct_delta(rev, p["营业收入(亿元)"])
    np_d = pct_delta(np_, p["归母净利润(亿元)"])
    ocf_d = pct_delta(ocf, p["经营现金流净额(亿元)"])
    dar_d = pct_delta(dar, p["资产负债率(%)"])
    roe_d = pct_delta(roe, p["ROE累计(%)"])
    bvps_d = pct_delta(bvps, p["每股净资产(元)"])
else:
    rev_d = np_d = ocf_d = dar_d = roe_d = bvps_d = ""

def fmt(v):
    return f"{v:.2f}"

# 替换卡片4 各 metric-value
# 营业收入
html = re.sub(r'(<div class="metric-name">营业收入</div>\s*<div class="metric-value dark">)[^<]*(<span class="metric-delta[^>]*>[^<]*</span>)?',
              lambda m: m.group(1) + f"{fmt(rev)} 亿" + (f'<span class="metric-delta down">{rev_d}</span>' if rev_d else ''),
              html, count=1)
# 归母净利润
html = re.sub(r'(<div class="metric-name">归母净利润（A股口径）</div>\s*<div class="metric-value dark">)[^<]*',
              lambda m: m.group(1) + f"≈{fmt(np_)} 亿元",
              html, count=1)
# 经营现金流
html = re.sub(r'(<div class="metric-name">经营性现金流净额</div>\s*<div class="metric-value dark">)[^<]*',
              lambda m: m.group(1) + f"{fmt(ocf)} 亿元",
              html, count=1)
# 资产负债率
html = re.sub(r'(<div class="metric-name">资产负债率</div>\s*<div class="metric-value green">)[^<]*',
              lambda m: m.group(1) + f"{fmt(dar)}%",
              html, count=1)
# ROE/分红率
html = re.sub(r'(<div class="metric-name">ROE / 分红率</div>\s*<div class="metric-value dark">)[^<]*',
              lambda m: m.group(1) + f"{fmt(roe)}% / 30~35%",
              html, count=1)
# 每股净资产
html = re.sub(r'(<div class="metric-name">每股净资产</div>\s*<div class="metric-value dark">)[^<]*',
              lambda m: m.group(1) + f"{fmt(bvps)} 元",
              html, count=1)

# 替换卡片4 各 metric 后的 mini sparkline（优先使用单季真实数据）
spark_map = [
    ("营业收入", "revenue"),
    ("归母净利润（A股口径）", "profit"),
    ("经营性现金流净额", "cash"),
    ("资产负债率", "debt"),
    ("ROE / 分红率", "roe"),
    ("每股净资产", "bps"),
]
for metric_name, key in spark_map:
    html = re.sub(
        rf'(<div class="metric-name">{re.escape(metric_name)}</div>\s*<div class="metric-value[^"]*">[^<]*(?:<span class="metric-delta[^>]*>[^<]*</span>)?</div>\s*</div>\s*)<svg class="spark"[^>]*>.*?</svg>',
        lambda m: m.group(1) + f'<img src="spark_{key}.svg" class="spark" alt="{metric_name}趋势">',
        html, count=1, flags=re.DOTALL)

# 备注：趋势图不再内嵌卡片4，统一在 HTML「长期趋势专区」区块展示（见 trend-box）。
# 此处仅校验趋势专区已引用两张图，不重复插入。
if 'trend_revenue_profit.svg' not in html or 'trend_ratios.svg' not in html:
    print("⚠️ 警告：HTML 长期趋势专区未引用趋势图 SVG，请检查。")

with open(HTML, "w", encoding="utf-8") as f:
    f.write(html)
print(f"✅ HTML 卡片4 已更新至最新报告期 {latest_date}；趋势图已归入『长期趋势专区』")
print(f"   营收 {fmt(rev)}亿({rev_d}) 归母 {fmt(np_)}亿 经营现金流 {fmt(ocf)}亿 负债率 {fmt(dar)}% ROE {fmt(roe)}% 每股净资产 {fmt(bvps)}元")
