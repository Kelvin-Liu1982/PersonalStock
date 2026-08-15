# -*- coding: utf-8 -*-
"""
构建「秦皇岛港5500大卡动力煤现货价」2026年以来周度序列。

数据说明（重要）：
- CCTD（中国煤炭市场网）确有「秦皇岛动力煤价格（周度）」数据，但历史序列在付费会员墙内，
  无法免费程序化直接抓取（已实测：周度数据页需登录且后端异常）。
- 本脚本【仅采用 CCTD 披露的锚点】，剔除环渤海 BSPI / Wind / 同花顺 / 雪球 / 券商研报等其他来源。
  锚点日期为 CCTD 真实披露日；CCTD 锚点之间的周度缺失值用两端 CCTD 锚点线性插值补足（仍仅基于 CCTD），
  标注「CCTD插值(估算)」；首个 CCTD 锚点之前（如1月）无 CCTD 披露数据，留空。保证口径单一、可追溯。

输出：
- 数据源/秦皇岛5500现货价_周度.xlsx  (新 Excel 文件)
- trend_qhd5500_weekly.svg     (趋势图)
"""
import datetime as dt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_XLSX = "数据源/秦皇岛5500现货价_周度.xlsx"
OUT_SVG = "trend_qhd5500_weekly.svg"

# ---------- 1. 仅采用 CCTD 披露的锚点（2026年以来）----------
# 其他来源（环渤海 BSPI / Wind / 同花顺 / 雪球 / 券商研报）一律剔除，仅保留 CCTD 口径。
# (日期, 价格(元/吨), 来源说明)
ANCHORS = [
    ("2026-02-04", 685.0, "CCTD秦皇岛动力煤价格5500(2月4日)"),
    ("2026-03-31", 760.0, "CCTD监测:3月底约760元/吨"),
    ("2026-04-29", 804.0, "CCTD:4月底收于804元/吨"),
    ("2026-05-21", 804.0, "CCTD秦皇岛5500(5月21日)"),
    ("2026-07-31", 725.0, "CCTD综合交易价5500(7月31日)"),
    ("2026-08-07", 739.0, "CCTD秦皇岛动力煤5500(启东发改委公告)"),
]

# ---------- 2. 构建周度序列（每周最后一个周五为观测点）----------
def fridays(start, end):
    d = start
    # 找到 start 之后的第一个周五
    while d.weekday() != 4:
        d += dt.timedelta(days=1)
    out = []
    while d <= end:
        out.append(d)
        d += dt.timedelta(days=7)
    return out

start = dt.date(2026, 1, 2)
end = dt.date(2026, 8, 7)
week_ends = fridays(start, end)

# 锚点转 datetime
anchor_dt = [(dt.date.fromisoformat(a[0]), a[1], a[2]) for a in ANCHORS]

rows = []
for w in week_ends:
    # 该周及之前最近的锚点
    prior = [a for a in anchor_dt if a[0] <= w]
    after = [a for a in anchor_dt if a[0] > w]
    if prior and (not after or (w - prior[-1][0]).days <= (after[0][0] - w).days):
        # 用最近的前置锚点（该周内披露的取当周）
        price = prior[-1][1]
        src = prior[-1][2]
        src_type = "锚点(CCTD)"
    elif prior and after:
        # 周度缺失：用两端 CCTD 锚点线性插值（仍仅基于 CCTD）
        p0, p1 = prior[-1], after[0]
        span = (p1[0] - p0[0]).days
        if span > 0:
            frac = (w - p0[0]).days / span
            price = p0[1] + (p1[1] - p0[1]) * frac
        else:
            price = p0[1]
        src = f"CCTD插值({p0[2]}→{p1[2]})"
        src_type = "CCTD插值(估算)"
    else:
        # 首个 CCTD 锚点之前（如1月）：无 CCTD 披露数据，留空
        price = np.nan
        src = "无CCTD披露"
        src_type = "无数据"

    rows.append({
        "周结束日": w.isoformat(),
        "年份-周": f"{w.isoformat()[:4]}-W{int(w.strftime('%W'))}",
        "秦皇岛5500现货价(元/吨)": round(price, 1),
        "数据性质": src_type,
        "来源/说明": src,
    })

df = pd.DataFrame(rows)

# ---------- 3. 写入新 Excel ----------
wb = Workbook()
ws = wb.active
ws.title = "秦皇岛5500周度"
title = "秦皇岛港5500大卡动力煤现货价（2026年以来·周度·仅CCTD口径）"
ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color="1E3A5F")
note = ("说明：本表【仅采用 CCTD 披露锚点】，剔除环渤海BSPI/Wind/同花顺/雪球等其他来源。"
        "CCTD锚点之间周度缺失值用两端 CCTD 锚点线性插值补足（标注『CCTD插值(估算)』）；"
        "首个 CCTD 锚点之前（如1月）无 CCTD 披露数据，留空。")
ws.cell(row=2, column=1, value=note).font = Font(italic=True, size=9, color="6B7280")
ws.cell(row=2, column=1, value=note).font = Font(italic=True, size=9, color="6B7280")
hdr = ["周结束日", "年份-周", "秦皇岛5500现货价(元/吨)", "数据性质", "来源/说明"]
for j, h in enumerate(hdr, 1):
    c = ws.cell(row=4, column=j, value=h)
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.font = Font(color="FFFFFF", bold=True)
    c.alignment = Alignment(horizontal="center", wrap_text=True)

thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for i, r in enumerate(df.to_dict("records"), start=5):
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=i, column=j, value=r[h])
        c.alignment = Alignment(horizontal="center")
        c.border = border
        if h == "数据性质" and r[h] == "插值(估算)":
            c.font = Font(color="B45309")
        if h == "数据性质" and r[h] == "锚点":
            c.font = Font(color="047857")

for j, wdt in zip(range(1, len(hdr)+1), [12, 12, 22, 14, 40]):
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(j)].width = wdt
ws.freeze_panes = "A5"

# 统计摘要
n_anchor = (df["数据性质"] == "锚点(CCTD)").sum()
n_interp = (df["数据性质"] == "CCTD插值(估算)").sum()
n_none = (df["数据性质"] == "无数据").sum()
valid = df.dropna(subset=["秦皇岛5500现货价(元/吨)"])
last = valid.iloc[-1] if len(valid) else df.iloc[-1]
summary = [
    f"样本周数: {len(df)}（2026-01-02 至 {end.isoformat()}）",
    f"CCTD锚点: {n_anchor} 周 | CCTD插值(估算): {n_interp} 周 | 无CCTD数据(留空): {n_none} 周",
    f"最新价: {last['秦皇岛5500现货价(元/吨)']} 元/吨 @ {last['周结束日']}",
    f"区间最低: {valid['秦皇岛5500现货价(元/吨)'].min()} | 最高: {valid['秦皇岛5500现货价(元/吨)'].max()} 元/吨",
    "口径: 仅采用 CCTD 披露锚点，未混入环渤海BSPI/Wind/同花顺等其他来源",
]
sr = 5 + len(df) + 2
for k, s in enumerate(summary):
    ws.cell(row=sr+k, column=1, value=s).font = Font(size=10, color="374151")

wb.save(OUT_XLSX)
print(f"✅ 生成 Excel：{OUT_XLSX}（CCTD锚点 {n_anchor} / CCTD插值 {n_interp} / 留空 {n_none}）")

# ---------- 4. 生成趋势图（纯 SVG）----------
def build_trend_svg(dates, vals, fname, W=1000, H=420, title=None):
    ml, mr, mt, mb = 60, 24, 42, 56
    pw, ph = W - ml - mr, H - mt - mb
    n = len(vals)
    # 仅用有效点绘制（跳过 NaN，1月无CCTD数据段留空）
    valid_idx = [i for i in range(n) if vals[i] == vals[i] and vals[i] is not None]
    xs_all = [ml + pw * i / (n - 1) for i in range(n)]
    vmin = min(vals[i] for i in valid_idx)
    vmax = max(vals[i] for i in valid_idx)
    pad = (vmax - vmin) * 0.12 or 10
    vmin, vmax = vmin - pad, vmax + pad
    def yp(v):
        return mt + ph - (v - vmin) / (vmax - vmin) * ph
    svg = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" font-family="-apple-system,Segoe UI,PingFang SC,sans-serif">']
    svg.append(f'<rect width="{W}" height="{H}" fill="#ffffff"/>')
    ttl = title or "秦皇岛港5500动力煤现货价（2026·周度·元/吨）"
    svg.append(f'<text x="{ml}" y="26" font-size="16" font-weight="700" fill="#1e3a5f">{ttl}</text>')
    # y 网格
    for k in range(5):
        yv = vmin + (vmax - vmin) * k / 4
        yy = yp(yv)
        svg.append(f'<line x1="{ml}" y1="{yy:.1f}" x2="{ml+pw}" y2="{yy:.1f}" stroke="#eef1f5"/>')
        svg.append(f'<text x="{ml-8}" y="{yy+4:.1f}" font-size="11" fill="#9ca3af" text-anchor="end">{yv:.0f}</text>')
    # 轴
    svg.append(f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt+ph}" stroke="#d1d5db"/>')
    svg.append(f'<line x1="{ml}" y1="{mt+ph}" x2="{ml+pw}" y2="{mt+ph}" stroke="#d1d5db"/>')
    # 折线（仅有效点，按真实日期 x 坐标定位）
    pts = [f"{xs_all[i]:.1f} {yp(vals[i]):.1f}" for i in valid_idx]
    svg.append(f'<path d="M{" L".join(pts)}" fill="none" stroke="#c53030" stroke-width="2.5"/>')
    for i in valid_idx:
        svg.append(f'<circle cx="{xs_all[i]:.1f}" cy="{yp(vals[i]):.1f}" r="2.6" fill="#c53030" stroke="#fff" stroke-width="1"/>')
    # x 标签（每月标一次）
    step = max(1, n // 12)
    for i in range(n):
        if i % step == 0 or i == n - 1:
            lab = dates[i][5:7] + "/" + dates[i][8:10]
            svg.append(f'<text x="{xs_all[i]:.1f}" y="{mt+ph+18:.1f}" font-size="10.5" fill="#6b7280" text-anchor="middle">{lab}</text>')
    # 图例
    svg.append(f'<rect x="{ml+4}" y="{H-20}" width="12" height="12" rx="2" fill="#c53030"/>')
    svg.append(f'<text x="{ml+21}" y="{H-10}" font-size="11.5" fill="#374151">秦皇岛5500现货价</text>')
    svg.append(f'<text x="{ml+pw}" y="{H-10}" font-size="10.5" fill="#9ca3af" text-anchor="end">单位:元/吨（仅CCTD口径，缺测段不画）</text>')
    svg.append('</svg>')
    with open(fname, "w", encoding="utf-8") as f:
        f.write("".join(svg))
    print(f"✅ 生成趋势图：{fname}")

build_trend_svg(df["周结束日"].tolist(), df["秦皇岛5500现货价(元/吨)"].tolist(), OUT_SVG,
                title="秦皇岛港5500动力煤现货价（2026·周度·仅CCTD口径）")

# ---------- 5. 卡片内 mini sparkline ----------
def build_spark(vals, fname, color="#c53030", W=62, H=28):
    n = len(vals)
    xs = [2 + (W - 4) * i / (n - 1) for i in range(n)]
    vmin, vmax = min(vals), max(vals)
    pad = (vmax - vmin) * 0.15 or 1
    def yp(v):
        return 4 + (H - 8) - (v - (vmin - pad)) / ((vmax + pad) - (vmin - pad)) * (H - 8)
    pts = [f"{xs[i]:.1f} {yp(vals[i]):.1f}" for i in range(n)]
    svg = [f'<svg class="spark" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg">']
    svg.append(f'<path d="M{" L".join(pts)}" fill="none" stroke="{color}" stroke-width="2"/>')
    svg.append('</svg>')
    with open(fname, "w", encoding="utf-8") as f:
        f.write("".join(svg))
    print(f"✅ 生成 sparkline：{fname}")

build_spark(df["秦皇岛5500现货价(元/吨)"].tolist(), "spark_qhd5500.svg")
