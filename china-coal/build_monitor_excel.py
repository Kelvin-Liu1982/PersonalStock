# -*- coding: utf-8 -*-
"""
中煤能源(601898.SH) 业务追踪监测数据 -> Excel 生成脚本
-----------------------------------------------------------------
数据来源：A股财报（新浪财经接口，akshare.stock_financial_report_sina）
sheet 规划：
  1) 利润表        —— 季报/年报原始数据（按报告日）
  2) 资产负债表    —— 季报/年报原始数据
  3) 现金流量表    —— 季报/年报原始数据
  4) 财务指标      —— 由三表派生的核心指标（营收/归母净利/经营现金流/资产负债率/每股净资产/ROE等）
  5) 产销与储量    —— 卡片1数据（年报口径，财务可印证部分 + 经营数据来源标注）
  6) 月度实时数据  —— 月频跟踪项模板（煤价/日耗/库存/产销月报等，留空待填）
  7) 数据字典      —— 各 sheet 来源与口径说明
"""
import akshare as ak
import pandas as pd
import numpy as np
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 注入带重试的会话，规避偶发 TLS EOF 抖动
_session = requests.Session()
_retry = Retry(total=5, backoff_factor=1.0,
               status_forcelist=[429, 500, 502, 503, 504],
               allowed_methods=frozenset(["GET", "POST"]))
for _p in ("https://", "http://"):
    _session.mount(_p, HTTPAdapter(max_retries=_retry))
ak.requests = _session
requests.get = _session.get  # akshare 内部部分直接引用 requests.get

STOCK = "sh601898"          # 中煤能源 A股
OUT = "数据源/中煤能源_业务追踪数据.xlsx"

# ---------- 1. 抓取财报三表 ----------
print("抓取利润表 ...")
income = ak.stock_financial_report_sina(stock=STOCK, symbol="利润表")
print("抓取资产负债表 ...")
balance = ak.stock_financial_report_sina(stock=STOCK, symbol="资产负债表")
print("抓取现金流量表 ...")
cashflow = ak.stock_financial_report_sina(stock=STOCK, symbol="现金流量表")

# 只保留合并口径（避免重复列），按报告日排序（旧->新）
def keep_consolidated(df):
    if "类型" in df.columns:
        df = df[df["类型"].astype(str).str.contains("合并", na=False)]
    df = df.copy()
    df["报告日"] = df["报告日"].astype(str)
    df = df.sort_values("报告日").reset_index(drop=True)
    return df

income = keep_consolidated(income)
balance = keep_consolidated(balance)
cashflow = keep_consolidated(cashflow)

# 报告日 -> 期间类型
def period_type(r):
    return "年度" if r.endswith("1231") else "季度"

# 仅保留 2020 年至今的报告期，聚焦近期且避免早期股本反推失真
KEEP_FROM = "20200101"
income = income[income["报告日"].astype(str) >= KEEP_FROM].reset_index(drop=True)
balance = balance[balance["报告日"].astype(str) >= KEEP_FROM].reset_index(drop=True)
cashflow = cashflow[cashflow["报告日"].astype(str) >= KEEP_FROM].reset_index(drop=True)

# ---------- 2. 派生财务指标 ----------
fin = pd.DataFrame()
fin["报告日"] = income["报告日"].astype(str)
fin["期间"] = fin["报告日"].map(period_type)
fin["营业收入(元)"] = income["营业收入"].astype(float)
fin["归母净利润(元)"] = income["归属于母公司所有者的净利润"].astype(float)
fin["利润总额(元)"] = income["利润总额"].astype(float)
fin["EPS基本(元)"] = income["基本每股收益"].astype(float)
fin["经营现金流净额(元)"] = cashflow["经营活动产生的现金流量净额"].astype(float)
fin["资产总计(元)"] = balance["资产总计"].astype(float)
fin["负债合计(元)"] = balance["负债合计"].astype(float)
fin["归母股东权益(元)"] = balance["归属于母公司股东权益合计"].astype(float)

# 股本（归母股东权益/EPS 反推较稳，这里用归母股东权益/每股净资产思路；先用公告股本 132.5亿股）
SHARES = 13_250_000_000.0
fin["资产负债率"] = fin["负债合计(元)"] / fin["资产总计(元)"]
fin["每股净资产(元)"] = fin["归母股东权益(元)"] / SHARES
fin["归母净利率"] = fin["归母净利润(元)"] / fin["营业收入(元)"]
# ROE(报告期累计，未年化)
fin["ROE(累计)"] = fin["归母净利润(元)"] / fin["归母股东权益(元)"]

# 单位换算为 亿元 / %
def to_yi(x):
    return round(x / 1e8, 2)
def to_pct(x):
    return round(x * 100, 2)

disp_fin = pd.DataFrame()
disp_fin["报告日"] = fin["报告日"]
disp_fin["期间"] = fin["期间"]
disp_fin["营业收入(亿元)"] = fin["营业收入(元)"].map(to_yi)
disp_fin["归母净利润(亿元)"] = fin["归母净利润(元)"].map(to_yi)
disp_fin["经营现金流净额(亿元)"] = fin["经营现金流净额(元)"].map(to_yi)
disp_fin["EPS基本(元)"] = fin["EPS基本(元)"].round(2)
disp_fin["每股净资产(元)"] = fin["每股净资产(元)"].round(2)
disp_fin["资产负债率(%)"] = fin["资产负债率"].map(to_pct)
disp_fin["归母净利率(%)"] = fin["归母净利率"].map(to_pct)
disp_fin["ROE累计(%)"] = fin["ROE(累计)"].map(to_pct)

# ---------- 3. 卡片1 产销与储量（年报口径，财务可印证 + 经营数据来源标注）----------
# 注：产量/销量/储量/产能为经营数据，A股标准财报三表不含，来源于公司年报"经营情况"
# 与月度经营数据公告；此处以卡片已有年报基线值录入，并标注来源，待月报/年报回填。
card1 = pd.DataFrame([
    ["自产商品煤产量", 1.351, "亿吨", "年报/月报", "公司披露", "FY2025 实际"],
    ["商品煤总销量", 2.56, "亿吨", "年报/月报", "公司披露", "FY2025 实际"],
    ["自产煤", 1.36, "亿吨", "年报分部", "公司披露", "FY2025 实际"],
    ["贸易煤", 1.20, "亿吨", "年报分部", "公司披露", "FY2025 实际"],
    ["在产矿井核定产能", 1.70, "亿吨", "年报/能源局核准", "公司+监管", "FY2025 在产"],
    ["剩余可采储量", 139.0, "亿吨", "年报储量章节", "公司披露(评估备案)", "FY2025"],
    ["在建-里必煤矿", 400, "万吨", "项目公告/核准", "公司+发改委", "在建"],
    ["在建-苇子沟煤矿", 240, "万吨", "项目公告/核准", "公司+发改委", "在建"],
    ["储产比(静态)", 102.9, "年", "=储量/自产产量", "派生", "139/1.351≈103年(卡片原写57年需核对)"],
], columns=["指标", "数值", "单位", "数据源", "披露主体", "口径说明"])

# ---------- 4. 月度实时数据模板 ----------
month_cols = ["年月", "秦皇岛5500现货(元/吨)", "沿海八省日耗(万吨)", "电厂库存天数",
              "自产商品煤产量(万吨)", "商品煤总销量(万吨)", "聚烯烃价格(元/吨)",
              "进口煤量(万吨)", "数据来源", "备注"]
month_rows = [
    ["2026-01", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", ""],
    ["2026-02", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", ""],
    ["2026-03", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", ""],
    ["2026-04", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", ""],
    ["2026-05", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", ""],
    ["2026-06", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", ""],
    ["2026-07", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", ""],
    ["2026-08", None, None, None, None, None, None, None, "CCTD/中电联/公司月报", "截至编制日"],
]
month_df = pd.DataFrame(month_rows, columns=month_cols)

# ---------- 5. 数据字典 ----------
dict_df = pd.DataFrame([
    ["利润表", "akshare 新浪财报-利润表(合并)", "季报+年报", "营业收入/归母净利润/EPS等"],
    ["资产负债表", "akshare 新浪财报-资产负债表(合并)", "季报+年报", "资产/负债/归母权益"],
    ["现金流量表", "akshare 新浪财报-现金流量表(合并)", "季报+年报", "经营/投资/筹资现金流"],
    ["财务指标", "由上述三表派生计算", "季报+年报", "资产负债率/每股净资产/ROE等"],
    ["产销与储量", "公司年报经营情况+月度经营公告+项目核准", "年报/月报/事件", "产量/销量/储量/产能(经营数据)"],
    ["月度实时数据", "CCTD/中电联/公司月报/统计局/海关", "月度", "煤价/日耗/库存/产销月报(待填)"],
], columns=["Sheet", "来源", "频率", "内容"])

# ---------- Excel 写出（含样式） ----------
wb = Workbook()
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1E3A5F")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_sheet(ws, df, title=None, num_fmt_cols=None):
    r0 = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
        r0 = 3
    # header
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=r0, column=j, value=str(col))
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    # body
    for i, (_, row) in enumerate(df.iterrows(), start=r0 + 1):
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            if isinstance(v, float) and np.isnan(v):
                v = None
            c = ws.cell(row=i, column=j, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            if num_fmt_cols and col in num_fmt_cols:
                c.number_format = "0.00"
    # width
    for j, col in enumerate(df.columns, start=1):
        maxlen = max([len(str(col))] + [len(str(row[col])) for _, row in df.iterrows()])
        ws.column_dimensions[get_column_letter(j)].width = min(max(maxlen + 2, 10), 42)
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1)

# Sheet1 利润表
ws = wb.active; ws.title = "利润表"
inc_disp = income[["报告日", "营业收入", "营业总成本", "营业利润", "利润总额",
                   "净利润", "归属于母公司所有者的净利润", "基本每股收益", "公告日期"]].copy()
for c in inc_disp.columns:
    if c != "报告日" and c != "公告日期":
        inc_disp[c] = inc_disp[c].astype(float).map(to_yi)
inc_disp.columns = ["报告日", "营业收入(亿元)", "营业总成本(亿元)", "营业利润(亿元)",
                    "利润总额(亿元)", "净利润(亿元)", "归母净利润(亿元)", "EPS基本(元)", "公告日期"]
write_sheet(ws, inc_disp, title="利润表（合并口径·单位:亿元）")

# Sheet2 资产负债表
ws = wb.create_sheet("资产负债表")
bal_disp = balance[["报告日", "资产总计", "负债合计", "流动负债合计", "非流动负债合计",
                    "归属于母公司股东权益合计", "公告日期"]].copy()
for c in bal_disp.columns:
    if c != "报告日" and c != "公告日期":
        bal_disp[c] = bal_disp[c].astype(float).map(to_yi)
bal_disp.columns = ["报告日", "资产总计(亿元)", "负债合计(亿元)", "流动负债(亿元)",
                   "非流动负债(亿元)", "归母股东权益(亿元)", "公告日期"]
write_sheet(ws, bal_disp, title="资产负债表（合并口径·单位:亿元）")

# Sheet3 现金流量表
ws = wb.create_sheet("现金流量表")
cf_disp = cashflow[["报告日", "经营活动产生的现金流量净额", "投资活动产生的现金流量净额",
                   "筹资活动产生的现金流量净额", "现金及现金等价物净增加额", "期末现金及现金等价物余额",
                   "公告日期"]].copy()
for c in cf_disp.columns:
    if c != "报告日" and c != "公告日期":
        cf_disp[c] = cf_disp[c].astype(float).map(to_yi)
cf_disp.columns = ["报告日", "经营现金流净额(亿元)", "投资现金流净额(亿元)", "筹资现金流净额(亿元)",
                 "现金净增加(亿元)", "期末现金余额(亿元)", "公告日期"]
write_sheet(ws, cf_disp, title="现金流量表（合并口径·单位:亿元）")

# Sheet4 财务指标
ws = wb.create_sheet("财务指标")
write_sheet(ws, disp_fin, title="核心财务指标（派生·单位:亿元/%）",
            num_fmt_cols=set(disp_fin.columns) - {"报告日", "期间"})

# Sheet5 产销与储量
ws = wb.create_sheet("产销与储量")
write_sheet(ws, card1, title="卡片1·煤炭产销与资源储量（年报基线）")

# Sheet6 月度实时数据
ws = wb.create_sheet("月度实时数据")
write_sheet(ws, month_df, title="月度高频跟踪数据（待回填·单位见列名）")

# Sheet7 数据字典
ws = wb.create_sheet("数据字典")
write_sheet(ws, dict_df, title="数据来源与口径字典")

wb.save(OUT)
print(f"\n✅ 已生成：{OUT}")
print(f"   sheet 数：{len(wb.sheetnames)} -> {wb.sheetnames}")
print(f"   利润表行数：{len(inc_disp)}，财务指标行数：{len(disp_fin)}")
print(f"   最新报告期：{disp_fin['报告日'].iloc[-1]} 营业收入 {disp_fin['营业收入(亿元)'].iloc[-1]}亿 / 归母 {disp_fin['归母净利润(亿元)'].iloc[-1]}亿")
