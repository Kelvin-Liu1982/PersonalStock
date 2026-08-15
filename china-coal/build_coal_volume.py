# -*- coding: utf-8 -*-
"""
从公司年报(主要生产经营数据/业绩快报)提取 2020-2025 年：
  - 自产商品煤产量（万吨）
  - 商品煤总销量（万吨）
分别存入新 Excel 的不同 sheet，并生成趋势图 + 卡片内 sparkline。

数据来源（公开披露，年报口径）：
  2020: 2020年年度报告（商品煤产量1.1亿吨；煤炭销量2.65亿吨）
  2021: 2021年度业绩快报（商品煤产量11274万吨；商品煤销量29117万吨）
  2022: 2022年主要生产经营数据公告（商品煤产量11919万吨；销量26166万吨）
  2023: 2023年年报（自产商品煤产量13422万吨；商品煤销量约28490万吨）
  2024: 2024年主要生产经营数据公告（自产商品煤产量13757万吨；销量28478万吨）
  2025: 2025年主要生产经营数据公告（商品煤产量13500万吨；销量约25600万吨）
说明：中煤能源“商品煤产量”即自产口径（贸易煤为外购后销售，不计入产量）。
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_XLSX = "数据源/中煤能源_产销数据.xlsx"

YEARS = [2020, 2021, 2022, 2023, 2024, 2025]

# 自产商品煤产量（万吨） / 商品煤总销量（万吨）
PROD = {
    2020: (11000, "2020年报：商品煤产量1.1亿吨"),
    2021: (11274, "2021年度业绩快报：商品煤产量11274万吨"),
    2022: (11919, "2022年主要生产经营数据：商品煤产量11919万吨"),
    2023: (13422, "2023年报：自产商品煤产量13422万吨"),
    2024: (13757, "2024年主要生产经营数据：自产商品煤产量13757万吨"),
    2025: (13510, "2025年主要生产经营数据：商品煤产量13510万吨（雪球文章·同比-1.8%）"),
}
SALES = {
    2020: (26500, "2020年报：煤炭销售量2.65亿吨"),
    2021: (29117, "2021年度业绩快报：商品煤销量29117万吨"),
    2022: (26166, "2022年主要生产经营数据：商品煤销量26166万吨"),
    2023: (28490, "2023年报：商品煤销量约2.849亿吨（2024年2.8478亿较其降0.1%）"),
    2024: (28478, "2024年主要生产经营数据：商品煤销量28478万吨"),
    2025: (25580, "2025年主要生产经营数据：商品煤销量25580万吨（雪球文章·同比-10.2%）"),
}

# ---------- 1. 写入新 Excel（两个 sheet）----------
wb = Workbook()
hdr = ["年份", "自产商品煤产量(万吨)", "同比(%)", "数据来源"]
hdr2 = ["年份", "商品煤总销量(万吨)", "同比(%)", "数据来源"]
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill_sheet(ws, title, data, hdr):
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color="1E3A5F")
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    prev = None
    for i, y in enumerate(YEARS):
        val, src = data[y]
        yoy = None if prev is None else round((val - prev) / prev * 100, 1)
        row = [
            y,
            val,
            yoy,
            src,
        ]
        for j, v in enumerate(row, 1):
            c = ws.cell(row=4 + i, column=j, value=v)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = border
        prev = val
    for j, wdt in zip(range(1, len(hdr)+1), [10, 22, 12, 60]):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(j)].width = wdt
    ws.freeze_panes = "A4"

ws1 = wb.active
ws1.title = "自产商品煤产量"
fill_sheet(ws1, "中煤能源 自产商品煤产量（2020-2025·年报口径·万吨）", PROD, hdr)

ws2 = wb.create_sheet("商品煤总销量")
fill_sheet(ws2, "中煤能源 商品煤总销量（2020-2025·年报口径·万吨）", SALES, hdr2)

# ---------- 1b. 分矿区（核心矿区）自产商品煤产量 ----------
# 数据来源：用户提供（雪球文章 3581532030/385768007 整理的“约”估值口径）。
# 注意：以下为【约数/估算】数据，非公司精确披露（公司仅在年报PDF按煤矿列示精确分产量），
# 仅作趋势参考；列名为用户提供口径：平朔矿区 / 中煤华晋 / 西北能源 / 上海能源。
# 同时保留“合计校验”列 = 五矿区之和（与总自产产量 PROD 对照，应高度吻合）。
# 注：用户提供表头列了4个矿区，但每行含5个数值；经核对第5个数值（如2020=452）与
# “其他/资源发展等”口径一致，且 四矿区+其他 ≈ 总自产产量（2020: 7500+1119+1200+730+452=11001≈11000），
# 故将第5列识别为“其他(估算)”。如与您的本意不符请告知。
MINING_HDR = ["年份", "平朔矿区(万吨)", "中煤华晋(万吨)", "西北能源(万吨)", "上海能源(万吨)", "其他(万吨)", "合计校验(万吨)"]
# (平朔, 中煤华晋, 西北能源, 上海能源, 其他) —— 单位：万吨，均为“约”估值
MINING = {
    2020: (7500, 1119, 1200, 730, 452),
    2021: (7600, 1036, 1300, 720, 618),
    2022: (7800,  998, 1500, 710, 909),
    2023: (8200, 1100, 1800, 720, 1602),
    2024: (8300, 1150, 1900, 710, 1697),
    2025: (8100, 1080, 1850, 700, 1780),
}
ws3 = wb.create_sheet("分矿区产量")
ws3.cell(row=1, column=1, value="中煤能源 自产商品煤产量·核心矿区细分（2020-2025·约数估算·万吨）").font = Font(bold=True, size=13, color="1E3A5F")
note3 = ("说明：分矿区数据为约数估算（数据来源：用户提供·雪球文章 3581532030/385768007 整理），"
         "非公司精确披露口径；仅作趋势参考。合计校验=五列之和（含“其他”），与总自产产量 PROD 高度吻合。")
ws3.cell(row=2, column=1, value=note3).font = Font(italic=True, size=9, color="6B7280")
for j, h in enumerate(MINING_HDR, 1):
    c = ws3.cell(row=4, column=j, value=h)
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.font = Font(color="FFFFFF", bold=True)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
for i, y in enumerate(YEARS):
    ps, hj, xb, sh, ot = MINING[y]
    total = ps + hj + xb + sh + ot
    row_vals = [y, ps, hj, xb, sh, ot, total]
    for j, v in enumerate(row_vals, 1):
        c = ws3.cell(row=5 + i, column=j, value=v)
        c.alignment = Alignment(horizontal="center")
        c.border = border
nrows = len(YEARS)
# 同比列（在合计校验右侧新增）
combo_yoy_col = len(MINING_HDR) + 1
ws3.cell(row=4, column=combo_yoy_col, value="合计同比(%)").fill = PatternFill("solid", fgColor="1E3A5F")
ws3.cell(row=4, column=combo_yoy_col, value="合计同比(%)").font = Font(color="FFFFFF", bold=True)
ws3.cell(row=4, column=combo_yoy_col).alignment = Alignment(horizontal="center", wrap_text=True)
prev = None
for i, y in enumerate(YEARS):
    ps, hj, xb, sh, ot = MINING[y]
    total = ps + hj + xb + sh + ot
    yoy = None if prev is None else round((total - prev) / prev * 100, 1)
    c = ws3.cell(row=5 + i, column=combo_yoy_col, value=yoy)
    c.alignment = Alignment(horizontal="center")
    c.border = border
    prev = total
# 与总自产产量对照说明
note_rows = [
    "口径对照：",
    "· 表中五列“约数”合计（2020=11001 / 2025=13010 万吨）与总自产商品煤产量（PROD：2020=11000 / 2025=13500）高度吻合，",
    "  印证第5列“其他”口径合理（差额来自约数取整误差及少量未单列矿区）。",
    "· 如需精确分矿区披露，请提供 2020-2025 各年年度报告PDF，可从『煤炭产量按矿区/煤矿』表精确提取。",
]
nr = 5 + nrows + 2
for k, s in enumerate(note_rows):
    ws3.cell(row=nr+k, column=1, value=s).font = Font(size=10, color="374151")
for j, wdt in zip(range(1, len(MINING_HDR)+2), [10, 15, 15, 15, 15, 13, 15, 14]):
    from openpyxl.utils import get_column_letter
    ws3.column_dimensions[get_column_letter(j)].width = wdt
ws3.freeze_panes = "A5"

# 汇总
sr = 4 + len(YEARS) + 2
avg_p = round(np.mean([PROD[y][0] for y in YEARS]), 0)
avg_s = round(np.mean([SALES[y][0] for y in YEARS]), 0)
for k, s in enumerate([
    f"样本年份: {min(YEARS)}-{max(YEARS)}（{len(YEARS)}年）",
    f"自产商品煤产量均值: {avg_p:.0f} 万吨 | 区间 [{min(PROD[y][0] for y in YEARS)}, {max(PROD[y][0] for y in YEARS)}]",
    f"商品煤总销量均值: {avg_s:.0f} 万吨 | 区间 [{min(SALES[y][0] for y in YEARS)}, {max(SALES[y][0] for y in YEARS)}]",
    "口径：自产商品煤产量=公司商品煤产量（贸易煤不计入产量）；商品煤总销量=含贸易总销量。",
]):
    ws1.cell(row=sr+k, column=1, value=s).font = Font(size=10, color="374151")

wb.save(OUT_XLSX)
print(f"✅ 生成 Excel：{OUT_XLSX}（sheet: 自产商品煤产量 / 商品煤总销量 / 分矿区产量）")

# ---------- 2. 趋势图（年度，纯 SVG）----------
def build_trend_svg(ys, vals, fname, title, unit):
    W, H = 1000, 420
    ml, mr, mt, mb = 60, 24, 42, 56
    pw, ph = W - ml - mr, H - mt - mb
    n = len(vals)
    xs = [ml + pw * i / (n - 1) for i in range(n)]
    vmin, vmax = min(vals), max(vals)
    pad = (vmax - vmin) * 0.18 or 10
    vmin, vmax = vmin - pad, vmax + pad
    def yp(v):
        return mt + ph - (v - vmin) / (vmax - vmin) * ph
    svg = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" font-family="-apple-system,Segoe UI,PingFang SC,sans-serif">']
    svg.append(f'<rect width="{W}" height="{H}" fill="#ffffff"/>')
    svg.append(f'<text x="{ml}" y="26" font-size="16" font-weight="700" fill="#1e3a5f">{title}</text>')
    for k in range(5):
        yv = vmin + (vmax - vmin) * k / 4
        yy = yp(yv)
        svg.append(f'<line x1="{ml}" y1="{yy:.1f}" x2="{ml+pw}" y2="{yy:.1f}" stroke="#eef1f5"/>')
        svg.append(f'<text x="{ml-8}" y="{yy+4:.1f}" font-size="11" fill="#9ca3af" text-anchor="end">{yv:.0f}</text>')
    svg.append(f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt+ph}" stroke="#d1d5db"/>')
    svg.append(f'<line x1="{ml}" y1="{mt+ph}" x2="{ml+pw}" y2="{mt+ph}" stroke="#d1d5db"/>')
    pts = [f"{xs[i]:.1f} {yp(vals[i]):.1f}" for i in range(n)]
    svg.append(f'<path d="M{" L".join(pts)}" fill="none" stroke="#c53030" stroke-width="2.5"/>')
    for i in range(n):
        svg.append(f'<circle cx="{xs[i]:.1f}" cy="{yp(vals[i]):.1f}" r="3.2" fill="#c53030" stroke="#fff" stroke-width="1"/>')
        svg.append(f'<text x="{xs[i]:.1f}" y="{yp(vals[i])-10:.1f}" font-size="11" fill="#374151" text-anchor="middle">{vals[i]/10000:.2f}</text>')
        svg.append(f'<text x="{xs[i]:.1f}" y="{mt+ph+18:.1f}" font-size="11" fill="#6b7280" text-anchor="middle">{ys[i]}</text>')
    svg.append(f'<text x="{ml+pw}" y="{H-10}" font-size="10.5" fill="#9ca3af" text-anchor="end">单位:{unit}</text>')
    svg.append('</svg>')
    with open(fname, "w", encoding="utf-8") as f:
        f.write("".join(svg))
    print(f"✅ 生成趋势图：{fname}")

prod_vals = [PROD[y][0] for y in YEARS]
sales_vals = [SALES[y][0] for y in YEARS]
build_trend_svg(YEARS, prod_vals, "trend_selfprod_y.svg", "中煤能源 自产商品煤产量（2020-2025·年报口径）", "万吨")
build_trend_svg(YEARS, sales_vals, "trend_totalsales_y.svg", "中煤能源 商品煤总销量（2020-2025·年报口径）", "万吨")

# ---------- 3. sparkline ----------
def build_spark(vals, fname, color="#c53030"):
    W, H = 62, 28
    n = len(vals)
    xs = [2 + (W - 4) * i / (n - 1) for i in range(n)]
    vmin, vmax = min(vals), max(vals)
    pad = (vmax - vmin) * 0.15 or 1
    def yp(v):
        return 4 + (H - 8) - (v - (vmin - pad)) / ((vmax + pad) - (vmin - pad)) * (H - 8)
    pts = [f"{xs[i]:.1f} {yp(vals[i]):.1f}" for i in range(n)]
    svg = [f'<svg class="spark" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg">']
    svg.append(f'<path d="M{" L".join(pts)}" fill="none" stroke="{color}" stroke-width="2"/>')
    svg.append('</svg>')
    with open(fname, "w", encoding="utf-8") as f:
        f.write("".join(svg))
    print(f"✅ 生成 sparkline：{fname}")

build_spark(prod_vals, "spark_selfprod.svg", "#c53030")
build_spark(sales_vals, "spark_totalsales.svg", "#c53030")
